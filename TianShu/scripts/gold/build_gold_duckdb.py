"""
构建 DuckDB Gold G0/G1/G2 表。

G0 公共维表 | G1 业务维表 | G2 明细事实表。
所有字段必须来自 Silver 或已明确标记的派生规则，并同步写入中文语义元数据。

dim_date 的日期范围动态计算：从所有 Silver 日期列中获取 min/max，
生成完整日期序列（当前覆盖 2012-01-01 ~ 2026-12-31）。
dim_violation_type 的罚款金额来源于官方数据字典 Excel。
"""
import argparse
import os
import sys
from pathlib import Path


PROJECT_ROOT = Path(__file__).resolve().parents[2]
QUALITY_DIR = PROJECT_ROOT / "scripts" / "quality"
sys.path.insert(0, str(QUALITY_DIR))

from harness_config import load_harness_config  # noqa: E402

try:
    import duckdb
except ImportError:
    duckdb = None

try:
    import openpyxl
except ImportError:
    openpyxl = None

# 官方违章代码字典 Excel 路径
VIOLATION_DICT_XLSX = Path(
    r"D:\ProgramData\Datawarehouse\纽约市城市交通"
    r"\2026财年纽约停车违章罚单开具情况"
    r"\Parking_Violations_Issued_Data_Dictionary.xlsx"
)


GOLD_DIMENSIONS = {
    "dim_date": {
        "zh": "日期维表",
        "columns": {
            "date_key": ("日期键", "主键"),
            "date": ("日期", "时间字段"),
            "year": ("年", "维度属性"),
            "quarter": ("季度", "维度属性"),
            "month": ("月", "维度属性"),
            "week": ("ISO周号", "维度属性"),
            "day_of_week": ("星期序号", "维度属性"),
            "day_of_week_name": ("星期名称", "维度属性"),
            "is_weekend": ("是否周末", "质量/分类标记"),
            "fiscal_year": ("NYC财年", "维度属性"),
        },
    },
    "dim_taxi_zone": {
        "zh": "出租车区域维表",
        "columns": {
            "location_id": ("出租车区域编号", "主键"),
            "borough": ("行政区", "维度属性"),
            "zone_name": ("区域名称", "维度属性"),
            "service_zone": ("服务区域", "维度属性"),
            "is_unknown_zone": ("是否未知区域", "质量标记"),
        },
    },
    "dim_vehicle": {
        "zh": "车辆维表",
        "columns": {
            "vehicle_key": ("车辆维表代理键", "主键"),
            "license_number": ("牌照编号", "业务键"),
            "license_type": ("牌照类型", "维度属性"),
            "license_status": ("牌照状态", "维度属性"),
            "owner_name": ("车主姓名或公司名", "维度属性"),
            "expiration_date": ("牌照到期日期", "时间字段"),
            "dmv_plate_number": ("DMV车牌号", "维度属性"),
            "vin": ("车辆识别码", "维度属性"),
            "vehicle_make": ("车辆品牌", "维度属性"),
            "vehicle_model": ("车辆型号", "维度属性"),
            "vehicle_year": ("车辆年份", "维度属性"),
            "fuel_type": ("燃料类型", "维度属性"),
            "wav_flag": ("无障碍车辆标志", "维度属性"),
            "stretch_limo": ("是否加长豪华轿车", "维度属性"),
            "medallion_type": ("Medallion类型", "维度属性"),
            "base_number": ("基地编号", "弱关联键"),
            "base_name": ("基地名称", "维度属性"),
            "base_type": ("基地类型", "维度属性"),
            "base_address": ("基地地址", "维度属性"),
            "agent_number": ("代理编号", "维度属性"),
            "agent_name": ("代理名称", "维度属性"),
            "insurance_carrier": ("保险公司名称", "维度属性"),
            "insurance_policy_number": ("保险单号", "维度属性"),
            "last_date_updated": ("最后更新日期", "时间字段"),
        },
    },
    "dim_driver": {
        "zh": "司机维表",
        "columns": {
            "driver_key": ("司机维表代理键", "主键"),
            "license_number": ("司机牌照号", "业务键"),
            "driver_name": ("司机姓名", "维度属性"),
            "driver_type": ("司机类型", "维度属性"),
            "status_code": ("状态码", "维度属性"),
            "status_description": ("状态描述", "维度属性"),
            "expiration_date": ("牌照到期日期", "时间字段"),
            "wav_trained": ("是否WAV培训", "维度属性"),
            "last_date_updated": ("最后更新日期", "时间字段"),
        },
    },
    "dim_base": {
        "zh": "基地维表",
        "columns": {
            "base_key": ("基地维表代理键", "主键"),
            "base_license_number": ("基地牌照号", "业务键"),
            "base_name": ("基地名称", "维度属性"),
            "dba": ("经营别名", "维度属性"),
            "base_type": ("基地类型", "维度属性"),
            "base_address": ("基地地址", "维度属性"),
        },
    },
    "dim_violation_type": {
        "zh": "违章类型维表",
        "columns": {
            "violation_code": ("违章代码", "主键"),
            "violation_description": ("违章描述", "维度属性"),
            "standard_fine_amount": ("标准罚款金额（曼哈顿核心区）", "金额字段"),
            "penalty_amount": ("滞纳金金额（当前无数据源）", "金额字段"),
            "source_status": ("金额数据来源", "审核状态"),
        },
    },
    "fact_trips": {
        "zh": "出行事实表",
        "columns": {
            "trip_id": ("行程代理主键", "主键"),
            "trip_source": ("行程来源类型", "退化维度"),
            "pickup_date_key": ("接客日期键", "维度外键"),
            "pickup_location_id": ("上车区域编号", "维度外键"),
            "dropoff_location_id": ("下车区域编号", "维度外键"),
            "base_no": ("派车基地编号", "弱关联键"),
            "pickup_at": ("接客时间", "时间字段"),
            "dropoff_at": ("送客时间", "时间字段"),
            "passenger_count": ("乘客人数", "度量字段"),
            "distance_miles": ("行程距离英里", "度量字段"),
            "fare_amount": ("基础车费", "金额字段"),
            "total_amount": ("总费用", "金额字段"),
            "tip_amount": ("小费", "金额字段"),
            "tolls_amount": ("通行费", "金额字段"),
            "driver_pay": ("司机净收入", "金额字段"),
            "is_time_anomaly": ("是否时间异常", "质量标记"),
            "is_location_missing": ("是否位置缺失", "质量标记"),
            "is_distance_outlier": ("是否距离异常", "质量标记"),
        },
    },
    "fact_parking_violations": {
        "zh": "停车罚单事实表",
        "columns": {
            "violation_id": ("罚单代理主键", "主键"),
            "summons_number": ("罚单编号", "候选键"),
            "issue_date_key": ("开票日期键", "维度外键"),
            "violation_code": ("违章代码", "维度外键"),
            "plate_id": ("车牌号", "退化维度"),
            "registration_state": ("注册州", "退化维度"),
            "plate_type": ("车牌类型", "退化维度"),
            "vehicle_body_type": ("车身类型", "退化维度"),
            "vehicle_make": ("车辆品牌", "退化维度"),
            "vehicle_year": ("车辆年份", "退化维度"),
            "violation_county": ("违章所在县", "空间属性"),
            "violation_precinct": ("违章管辖区", "空间属性"),
            "issuing_agency": ("开票机构", "退化维度"),
            "feet_from_curb": ("距路缘英尺数", "度量字段"),
            "fiscal_year": ("财年", "时间属性"),
            "standard_fine_amount": ("标准罚款金额", "金额字段"),
            "fine_source_status": ("罚款金额来源状态", "审核状态"),
            "is_duplicate_summons": ("是否重复罚单", "质量标记"),
        },
    },
    "fact_tif_payments": {
        "zh": "TIF支付事实表",
        "columns": {
            "payment_id": ("支付代理主键", "主键"),
            "license_number": ("牌照号", "弱关联键"),
            "agent_number": ("代理编号", "退化维度"),
            "payment_date_key": ("支付日期键", "维度外键"),
            "hackup_payment_amount": ("改装支付金额", "金额字段"),
            "operational_payment_amount": ("运营支付金额", "金额字段"),
            "total_payment_amount": ("总支付金额", "金额字段"),
            "is_duplicate_key": ("是否复合键重复", "质量标记"),
        },
    },
    "fact_driver_applications": {
        "zh": "司机申请事实表",
        "columns": {
            "application_id": ("申请代理主键", "主键"),
            "app_no": ("申请编号", "业务键"),
            "application_type": ("申请类型", "退化维度"),
            "app_date_key": ("申请日期键", "维度外键"),
            "status": ("审批状态", "状态字段"),
            "drug_test": ("药检状态", "状态字段"),
            "wav_course": ("WAV培训状态", "状态字段"),
            "defensive_driving": ("防御性驾驶状态", "状态字段"),
            "driver_exam": ("司机考试状态", "状态字段"),
            "last_updated": ("最后更新日期", "时间字段"),
        },
    },
    "fact_crashes": {
        "zh": "事故事实表",
        "columns": {
            "crash_id": ("事故代理主键", "主键"),
            "collision_id": ("事故编号", "业务键"),
            "crash_date_key": ("事故日期键", "维度外键"),
            "borough": ("行政区", "退化维度"),
            "zip_code": ("邮政编码", "退化维度"),
            "latitude": ("纬度", "空间字段"),
            "longitude": ("经度", "空间字段"),
            "persons_injured": ("受伤总人数", "度量字段"),
            "persons_killed": ("死亡总人数", "度量字段"),
            "pedestrians_injured": ("行人受伤数", "度量字段"),
            "pedestrians_killed": ("行人死亡数", "度量字段"),
            "cyclist_injured": ("骑行者受伤数", "度量字段"),
            "cyclist_killed": ("骑行者死亡数", "度量字段"),
            "motorist_injured": ("驾驶员受伤数", "度量字段"),
            "motorist_killed": ("驾驶员死亡数", "度量字段"),
            "vehicle_type_1": ("涉事车辆1类型", "退化维度"),
            "vehicle_type_2": ("涉事车辆2类型", "退化维度"),
            "contributing_factor_1": ("车辆1事故因素", "退化维度"),
            "contributing_factor_2": ("车辆2事故因素", "退化维度"),
            "is_location_missing": ("是否位置缺失", "质量标记"),
        },
    },
    "fact_crash_persons": {
        "zh": "事故人员事实表",
        "columns": {
            "crash_person_id": ("事故人员代理主键", "主键"),
            "unique_id": ("人员记录编号", "业务键"),
            "collision_id": ("事故编号", "关联键"),
            "crash_date_key": ("事故日期键", "维度外键"),
            "person_type": ("人员类型", "退化维度"),
            "person_injury": ("伤害程度", "退化维度"),
            "person_sex": ("性别", "退化维度"),
            "person_age": ("年龄", "度量字段"),
            "ped_role": ("行人角色", "退化维度"),
            "ejection": ("是否弹出", "退化维度"),
            "emotional_status": ("情绪状态", "退化维度"),
            "bodily_injury": ("身体伤害", "退化维度"),
            "position_in_vehicle": ("车内位置", "退化维度"),
            "safety_equipment": ("安全设备", "退化维度"),
            "is_orphan_record": ("是否孤立记录", "质量标记"),
            "is_age_anomaly": ("是否年龄异常", "质量标记"),
        },
    },
}

GOLD_DIMENSIONS.update(
    {
        "dws_daily_trip_summary": {
            "zh": "每日出行汇总表",
            "columns": {
                "date_key": ("日期键", "维度外键"),
                "trip_date": ("出行日期", "时间字段"),
                "trip_count": ("行程量", "指标字段"),
                "total_passenger_count": ("乘客总人数", "指标字段"),
                "total_distance_miles": ("总行程距离英里", "指标字段"),
                "total_fare_amount": ("基础车费总额", "金额指标"),
                "avg_distance_miles": ("平均行程距离英里", "指标字段"),
                "time_anomaly_count": ("时间异常行程数", "质量指标"),
                "location_missing_count": ("位置缺失行程数", "质量指标"),
                "distance_outlier_count": ("距离异常行程数", "质量指标"),
            },
        },
        "dws_zone_trip_summary": {
            "zh": "区域出行汇总表",
            "columns": {
                "pickup_location_id": ("上车区域编号", "维度外键"),
                "borough": ("行政区", "维度属性"),
                "zone_name": ("区域名称", "维度属性"),
                "trip_count": ("行程量", "指标字段"),
                "total_distance_miles": ("总行程距离英里", "指标字段"),
                "total_fare_amount": ("基础车费总额", "金额指标"),
                "avg_distance_miles": ("平均行程距离英里", "指标字段"),
                "location_missing_count": ("位置缺失行程数", "质量指标"),
            },
        },
        "dws_daily_parking_summary": {
            "zh": "每日罚单汇总表",
            "columns": {
                "date_key": ("日期键", "维度外键"),
                "issue_date": ("开票日期", "时间字段"),
                "violation_count": ("罚单量", "指标字段"),
                "unique_plate_count": ("唯一车牌数", "指标字段"),
                "standard_fine_total": ("标准罚款总额", "金额指标"),
                "standard_fine_covered_count": ("有标准罚款金额罚单数", "质量指标"),
                "duplicate_summons_count": ("重复罚单数", "质量指标"),
            },
        },
        "dws_daily_crash_summary": {
            "zh": "每日事故汇总表",
            "columns": {
                "date_key": ("日期键", "维度外键"),
                "crash_date": ("事故日期", "时间字段"),
                "crash_count": ("事故量", "指标字段"),
                "persons_injured": ("受伤人数", "指标字段"),
                "persons_killed": ("死亡人数", "指标字段"),
                "pedestrians_injured": ("行人受伤数", "指标字段"),
                "pedestrians_killed": ("行人死亡数", "指标字段"),
                "cyclist_injured": ("骑行者受伤数", "指标字段"),
                "cyclist_killed": ("骑行者死亡数", "指标字段"),
                "motorist_injured": ("机动车人员受伤数", "指标字段"),
                "motorist_killed": ("机动车人员死亡数", "指标字段"),
                "location_missing_count": ("位置缺失事故数", "质量指标"),
            },
        },
    }
)


def connect_duckdb(db_path: Path):
    """连接 DuckDB，失败时给出明确错误"""
    if duckdb is None:
        raise RuntimeError("duckdb 未安装，无法构建 Gold 表")
    return duckdb.connect(str(db_path), read_only=False)


def ensure_meta_tables(conn) -> None:
    """确保中文语义元数据表存在"""
    conn.execute("CREATE SCHEMA IF NOT EXISTS meta")
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS meta.table_comments (
            table_schema VARCHAR,
            table_name VARCHAR,
            table_name_zh VARCHAR,
            updated_at VARCHAR
        )
        """
    )


def ensure_semantic_tables(conn) -> None:
    """确保 Gold 中文语义层表存在"""
    conn.execute("CREATE SCHEMA IF NOT EXISTS meta")
    conn.execute(
        """
        CREATE OR REPLACE TABLE meta.metric_definitions (
            metric_name VARCHAR,
            metric_name_zh VARCHAR,
            source_table VARCHAR,
            calculation_sql VARCHAR,
            time_key VARCHAR,
            business_meaning VARCHAR,
            audit_status VARCHAR
        )
        """
    )
    conn.execute(
        """
        CREATE OR REPLACE TABLE meta.semantic_dimensions (
            dimension_name VARCHAR,
            dimension_name_zh VARCHAR,
            source_table VARCHAR,
            source_column VARCHAR,
            business_meaning VARCHAR
        )
        """
    )
    conn.execute(
        """
        CREATE OR REPLACE TABLE meta.semantic_query_templates (
            question_zh VARCHAR,
            intent_name VARCHAR,
            recommended_table VARCHAR,
            metric_names VARCHAR,
            sql_template VARCHAR,
            caution VARCHAR
        )
        """
    )
    conn.execute(
        """
        CREATE OR REPLACE TABLE meta.business_terms (
            term VARCHAR,
            term_zh VARCHAR,
            definition_zh VARCHAR,
            source VARCHAR
        )
        """
    )


def write_semantic_layer(conn) -> None:
    """写入面向中文问数的指标、维度、模板和术语"""
    ensure_semantic_tables(conn)
    conn.executemany(
        "INSERT INTO meta.metric_definitions VALUES (?, ?, ?, ?, ?, ?, ?)",
        [
            ("trip_count", "行程量", "gold.dws_daily_trip_summary", "sum(trip_count)", "date_key", "统计出行事实表中的行程记录数", "approved"),
            ("total_fare_amount", "基础车费总额", "gold.dws_daily_trip_summary", "sum(total_fare_amount)", "date_key", "统计基础车费金额，不等同于乘客总支付金额", "approved"),
            ("avg_distance_miles", "平均行程距离", "gold.dws_daily_trip_summary", "avg(avg_distance_miles)", "date_key", "按日或区域观察平均行程距离", "approved"),
            ("parking_violation_count", "罚单量", "gold.dws_daily_parking_summary", "sum(violation_count)", "date_key", "统计停车违章罚单数量", "approved"),
            ("standard_fine_total", "标准罚款总额", "gold.dws_daily_parking_summary", "sum(standard_fine_total)", "date_key", "按官方标准罚款金额估算，不代表实际收款", "approved"),
            ("crash_count", "事故量", "gold.dws_daily_crash_summary", "sum(crash_count)", "date_key", "统计机动车碰撞事故数量", "approved"),
            ("persons_injured", "受伤人数", "gold.dws_daily_crash_summary", "sum(persons_injured)", "date_key", "统计事故涉及受伤人数", "approved"),
            ("persons_killed", "死亡人数", "gold.dws_daily_crash_summary", "sum(persons_killed)", "date_key", "统计事故涉及死亡人数", "approved"),
        ],
    )
    conn.executemany(
        "INSERT INTO meta.semantic_dimensions VALUES (?, ?, ?, ?, ?)",
        [
            ("date", "日期", "gold.dim_date", "date", "按自然日、月、季度、财年分析"),
            ("taxi_zone", "出租车区域", "gold.dim_taxi_zone", "zone_name", "按 TLC 出租车区域分析上车和下车热度"),
            ("borough", "行政区", "gold.dim_taxi_zone", "borough", "按纽约市行政区分析空间分布"),
            ("trip_source", "行程来源类型", "gold.fact_trips", "trip_source", "区分 yellow、green、fhv、fhvhv 行程"),
            ("violation_code", "违章代码", "gold.dim_violation_type", "violation_code", "按官方违章代码分析罚单"),
            ("person_type", "事故人员类型", "gold.fact_crash_persons", "person_type", "按驾驶员、乘客、行人、骑行者等分析事故人员"),
        ],
    )
    conn.executemany(
        "INSERT INTO meta.semantic_query_templates VALUES (?, ?, ?, ?, ?, ?)",
        [
            ("2026 年 Q1 每天有多少行程？", "daily_trip_count", "gold.dws_daily_trip_summary", "trip_count", "SELECT trip_date, trip_count FROM gold.dws_daily_trip_summary ORDER BY trip_date", "行程日期使用 pickup_date_key"),
            ("哪个区域上车量最高？", "top_pickup_zone", "gold.dws_zone_trip_summary", "trip_count", "SELECT borough, zone_name, trip_count FROM gold.dws_zone_trip_summary ORDER BY trip_count DESC LIMIT 20", "只统计 pickup_location_id 非空且可关联区域的行程"),
            ("每天停车罚单数量是多少？", "daily_parking_count", "gold.dws_daily_parking_summary", "parking_violation_count", "SELECT issue_date, violation_count FROM gold.dws_daily_parking_summary ORDER BY issue_date", "日期使用罚单开票日期 issue_date_key"),
            ("标准罚款金额最高的日期是哪天？", "daily_standard_fine", "gold.dws_daily_parking_summary", "standard_fine_total", "SELECT issue_date, standard_fine_total FROM gold.dws_daily_parking_summary ORDER BY standard_fine_total DESC LIMIT 20", "标准罚款金额不是实际缴款金额"),
            ("每天事故数量和死亡人数趋势如何？", "daily_crash_trend", "gold.dws_daily_crash_summary", "crash_count,persons_killed", "SELECT crash_date, crash_count, persons_killed FROM gold.dws_daily_crash_summary ORDER BY crash_date", "事故日期来自 crash_date_key"),
            ("每天事故受伤人数是多少？", "daily_injury_trend", "gold.dws_daily_crash_summary", "persons_injured", "SELECT crash_date, persons_injured FROM gold.dws_daily_crash_summary ORDER BY crash_date", "使用事故级受伤人数汇总，不用人员表重复计算"),
        ],
    )
    conn.executemany(
        "INSERT INTO meta.business_terms VALUES (?, ?, ?, ?)",
        [
            ("FHV", "网约出租车辆", "For-Hire Vehicle，纽约 TLC 语境中的预约出租车辆", "TLC"),
            ("HVFHV", "高容量网约出租车辆", "High Volume For-Hire Vehicle，Uber/Lyft 等高容量平台车辆类别", "TLC"),
            ("TIF", "出租车改善基金", "Taxi Improvement Fund，用于无障碍车辆改装和运营相关支付", "TLC"),
            ("WAV", "无障碍车辆", "Wheelchair Accessible Vehicle，可服务轮椅乘客的车辆", "TLC"),
            ("standard_fine_amount", "标准罚款金额", "来自官方违章代码字典的标准罚款金额，不代表实际收款", "NYC Open Data"),
        ],
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS meta.column_comments (
            table_schema VARCHAR,
            table_name VARCHAR,
            column_name VARCHAR,
            column_name_zh VARCHAR,
            column_role_zh VARCHAR,
            updated_at VARCHAR
        )
        """
    )


def load_violation_fines(conn) -> None:
    """将官方违章代码 Excel 加载到 DuckDB 临时表，供 dim_violation_type 建表使用"""
    if openpyxl is None:
        print("[WARN] openpyxl 未安装，dim_violation_type 罚款金额将保持 NULL")
        return

    if not VIOLATION_DICT_XLSX.exists():
        print(f"[WARN] 违章代码字典 Excel 不存在: {VIOLATION_DICT_XLSX}")
        return

    wb = openpyxl.load_workbook(VIOLATION_DICT_XLSX, data_only=True)
    ws = wb["Parking Violation Codes"]

    rows: list[tuple[int, str, int, int]] = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        code = row[0]
        desc = row[1]
        manhattan_fine = row[2]
        other_fine = row[3]
        if code is None or str(code).strip() == "":
            continue
        try:
            code_int = int(code)
        except (ValueError, TypeError):
            continue
        rows.append((
            code_int,
            str(desc).strip() if desc else "",
            int(manhattan_fine) if manhattan_fine is not None else 0,
            int(other_fine) if other_fine is not None else 0,
        ))

    wb.close()

    conn.execute("DROP TABLE IF EXISTS tmp_violation_fines")
    conn.execute(
        """
        CREATE TEMP TABLE tmp_violation_fines (
            violation_code_int INTEGER,
            violation_description_from_dict VARCHAR,
            manhattan_fine INTEGER,
            other_areas_fine INTEGER
        )
        """
    )
    conn.executemany(
        "INSERT INTO tmp_violation_fines VALUES (?, ?, ?, ?)", rows
    )
    print(f"已加载 {len(rows)} 条违章代码罚款标准（来源：官方数据字典 Excel）")


def replace_gold_comments(conn, table_names: list[str]) -> None:
    """重建本批 Gold 表对应的中文注释，避免残留旧字段"""
    placeholders = ", ".join(["?"] * len(table_names))
    conn.execute(
        f"DELETE FROM meta.table_comments WHERE table_schema = 'gold' AND table_name IN ({placeholders})",
        table_names,
    )
    conn.execute(
        f"DELETE FROM meta.column_comments WHERE table_schema = 'gold' AND table_name IN ({placeholders})",
        table_names,
    )


def write_gold_comments(conn, table_names: list[str]) -> None:
    """写入 Gold 表和字段中文注释"""
    timestamp_sql = "strftime(now(), '%Y-%m-%d %H:%M:%S')"
    for table_name in table_names:
        definition = GOLD_DIMENSIONS[table_name]
        conn.execute(
            f"""
            INSERT INTO meta.table_comments
            SELECT 'gold', ?, ?, {timestamp_sql}
            """,
            [table_name, definition["zh"]],
        )
        for column_name, (column_name_zh, column_role_zh) in definition["columns"].items():
            conn.execute(
                f"""
                INSERT INTO meta.column_comments
                SELECT 'gold', ?, ?, ?, ?, {timestamp_sql}
                """,
                [table_name, column_name, column_name_zh, column_role_zh],
            )


def build_g0(conn) -> list[str]:
    """构建 G0 公共维表。dim_date 从 Silver 实际日期范围动态生成。"""
    conn.execute("CREATE SCHEMA IF NOT EXISTS gold")

    # ── dim_date：从所有 Silver 日期列动态生成完整日期范围 ──
    # 避免硬编码固定区间，确保覆盖任何新导入的数据集
    conn.execute("DROP TABLE IF EXISTS gold.dim_date")
    conn.execute(
        """
        CREATE TABLE gold.dim_date AS
        WITH date_bounds AS (
            -- 每表独立取 min/max，避免 UNION 全表扫描
            SELECT 'trip' AS src, min(pickup_at)::DATE AS d_min, max(pickup_at)::DATE AS d_max
                FROM silver.trip_detail WHERE pickup_at IS NOT NULL
            UNION ALL
            SELECT 'parking', min(issue_date), max(issue_date)
                FROM silver.parking_violation_detail
                WHERE issue_date IS NOT NULL
                  AND issue_date >= '2000-01-01'
                  AND issue_date <= '2027-12-31'
            UNION ALL
            SELECT 'tif', min(payment_date), max(payment_date)
                FROM silver.tif_payment_detail WHERE payment_date IS NOT NULL
            UNION ALL
            SELECT 'driver_app', min(app_date), max(app_date)
                FROM silver.driver_application_detail WHERE app_date IS NOT NULL
            UNION ALL
            SELECT 'crash', min(crash_at)::DATE, max(crash_at)::DATE
                FROM silver.crash_detail WHERE crash_at IS NOT NULL
            UNION ALL
            SELECT 'crash_person', min(crash_date), max(crash_date)
                FROM silver.crash_person_detail WHERE crash_date IS NOT NULL
        ),
        global_bounds AS (
            SELECT
                CAST(date_trunc('year', min(d_min)) AS DATE) AS start_date,
                CAST(date_trunc('year', max(d_max) + INTERVAL 1 YEAR) - INTERVAL 1 DAY AS DATE) AS end_date
            FROM date_bounds
        ),
        date_range AS (
            SELECT UNNEST(generate_series(
                (SELECT start_date FROM global_bounds),
                (SELECT end_date FROM global_bounds),
                INTERVAL 1 DAY
            )) AS date
        )
        SELECT
            strftime(date, '%Y%m%d')::INTEGER AS date_key,
            date,
            year(date) AS year,
            quarter(date) AS quarter,
            month(date) AS month,
            week(date) AS week,
            dayofweek(date) AS day_of_week,
            CASE dayofweek(date)
                WHEN 1 THEN '周一' WHEN 2 THEN '周二' WHEN 3 THEN '周三'
                WHEN 4 THEN '周四' WHEN 5 THEN '周五' WHEN 6 THEN '周六'
                WHEN 0 THEN '周日'
            END AS day_of_week_name,
            CASE WHEN dayofweek(date) IN (0, 6) THEN true ELSE false END AS is_weekend,
            CASE
                WHEN month(date) >= 7
                    THEN year(date) + 1
                ELSE year(date)
            END AS fiscal_year
        FROM date_range
        """
    )

    # ── dim_taxi_zone ──
    conn.execute("DROP TABLE IF EXISTS gold.dim_taxi_zone")
    conn.execute(
        """
        CREATE TABLE gold.dim_taxi_zone AS
        SELECT
            location_id,
            borough,
            zone_name,
            service_zone,
            is_unknown_zone
        FROM silver.taxi_zone
        """
    )
    return ["dim_date", "dim_taxi_zone"]


def build_g1(conn) -> list[str]:
    """构建 G1 业务维表"""
    conn.execute("DROP TABLE IF EXISTS gold.dim_vehicle")
    conn.execute(
        """
        CREATE TABLE gold.dim_vehicle AS
        SELECT
            vehicle_id AS vehicle_key,
            license_number,
            license_type,
            license_status,
            owner_name,
            expiration_date,
            dmv_plate_number,
            vin,
            vehicle_make,
            vehicle_model,
            vehicle_year,
            fuel_type,
            wav_flag,
            stretch_limo,
            medallion_type,
            base_number,
            base_name,
            base_type,
            base_address,
            agent_number,
            agent_name,
            insurance_carrier,
            insurance_policy_number,
            last_date_updated
        FROM silver.vehicle_detail
        """
    )

    conn.execute("DROP TABLE IF EXISTS gold.dim_driver")
    conn.execute(
        """
        CREATE TABLE gold.dim_driver AS
        SELECT
            driver_id AS driver_key,
            license_number,
            driver_name,
            driver_type,
            status_code,
            status_description,
            expiration_date,
            wav_trained,
            last_date_updated
        FROM silver.driver_detail
        """
    )

    conn.execute("DROP TABLE IF EXISTS gold.dim_base")
    conn.execute(
        """
        CREATE TABLE gold.dim_base AS
        WITH latest_base AS (
            SELECT
                base_license_number,
                base_name,
                dba,
                row_number() OVER (
                    PARTITION BY base_license_number
                    ORDER BY year DESC NULLS LAST, month DESC NULLS LAST, base_detail_id DESC
                ) AS rn
            FROM silver.base_detail
            WHERE base_license_number IS NOT NULL
        ),
        vehicle_base AS (
            SELECT
                base_number,
                max(base_type) AS base_type,
                max(base_address) AS base_address
            FROM silver.vehicle_detail
            WHERE base_number IS NOT NULL
            GROUP BY base_number
        ),
        base_one_row AS (
            SELECT
                lb.base_license_number,
                lb.base_name,
                lb.dba,
                vb.base_type,
                vb.base_address
            FROM latest_base lb
            LEFT JOIN vehicle_base vb
              ON vb.base_number = lb.base_license_number
            WHERE lb.rn = 1
        )
        SELECT
            row_number() OVER (ORDER BY base_license_number) AS base_key,
            base_license_number,
            base_name,
            dba,
            base_type,
            base_address
        FROM base_one_row
        """
    )

    conn.execute("DROP TABLE IF EXISTS gold.dim_violation_type")

    # 从官方 Excel 加载违章代码罚款标准
    load_violation_fines(conn)

    conn.execute(
        """
        CREATE TABLE gold.dim_violation_type AS
        WITH silver_codes AS (
            SELECT
                violation_code,
                max(violation_description) AS violation_description
            FROM silver.parking_violation_detail
            WHERE violation_code IS NOT NULL
            GROUP BY violation_code
        )
        SELECT
            s.violation_code,
            COALESCE(d.violation_description_from_dict, s.violation_description)
                AS violation_description,
            CAST(d.manhattan_fine AS DECIMAL(12,2)) AS standard_fine_amount,
            CAST(NULL AS DECIMAL(12,2)) AS penalty_amount,
            CASE
                WHEN d.violation_code_int IS NOT NULL THEN 'from_official_dictionary'
                ELSE 'missing_from_dictionary'
            END AS source_status
        FROM silver_codes s
        LEFT JOIN tmp_violation_fines d
          ON TRY_CAST(s.violation_code AS INTEGER) = d.violation_code_int
        """
    )
    return ["dim_vehicle", "dim_driver", "dim_base", "dim_violation_type"]


def build_g2(conn) -> list[str]:
    """构建 G2 明细事实表"""
    conn.execute("DROP TABLE IF EXISTS gold.fact_trips")
    conn.execute(
        """
        CREATE TABLE gold.fact_trips AS
        SELECT
            trip_id,
            trip_source,
            CASE
                WHEN pickup_date IS NOT NULL
                THEN strftime(pickup_date, '%Y%m%d')::INTEGER
                ELSE NULL
            END AS pickup_date_key,
            pickup_location_id,
            dropoff_location_id,
            base_no,
            pickup_at,
            dropoff_at,
            passenger_count,
            distance_miles,
            fare_amount,
            total_amount,
            tip_amount,
            tolls_amount,
            driver_pay,
            is_time_anomaly,
            is_location_missing,
            is_distance_outlier
        FROM silver.trip_detail
        """
    )

    conn.execute("DROP TABLE IF EXISTS gold.fact_parking_violations")
    conn.execute(
        """
        CREATE TABLE gold.fact_parking_violations AS
        SELECT
            p.violation_id,
            p.summons_number,
            CASE
                WHEN p.issue_date IS NOT NULL
                THEN strftime(p.issue_date, '%Y%m%d')::INTEGER
                ELSE NULL
            END AS issue_date_key,
            p.violation_code,
            p.plate_id,
            p.registration_state,
            p.plate_type,
            p.vehicle_body_type,
            p.vehicle_make,
            p.vehicle_year,
            p.violation_county,
            p.violation_precinct,
            p.issuing_agency,
            p.feet_from_curb,
            p.fiscal_year,
            d.standard_fine_amount,
            d.source_status AS fine_source_status,
            p.is_duplicate_summons
        FROM silver.parking_violation_detail p
        LEFT JOIN gold.dim_violation_type d
          ON d.violation_code = p.violation_code
        """
    )

    conn.execute("DROP TABLE IF EXISTS gold.fact_tif_payments")
    conn.execute(
        """
        CREATE TABLE gold.fact_tif_payments AS
        SELECT
            payment_id,
            license_number,
            agent_number,
            CASE
                WHEN payment_date IS NOT NULL
                THEN strftime(payment_date, '%Y%m%d')::INTEGER
                ELSE NULL
            END AS payment_date_key,
            hackup_payment_amount,
            operational_payment_amount,
            total_payment_amount,
            is_duplicate_key
        FROM silver.tif_payment_detail
        """
    )

    conn.execute("DROP TABLE IF EXISTS gold.fact_driver_applications")
    conn.execute(
        """
        CREATE TABLE gold.fact_driver_applications AS
        SELECT
            application_id,
            app_no,
            application_type,
            CASE
                WHEN app_date IS NOT NULL
                THEN strftime(app_date, '%Y%m%d')::INTEGER
                ELSE NULL
            END AS app_date_key,
            status,
            drug_test,
            wav_course,
            defensive_driving,
            driver_exam,
            last_updated
        FROM silver.driver_application_detail
        """
    )

    conn.execute("DROP TABLE IF EXISTS gold.fact_crashes")
    conn.execute(
        """
        CREATE TABLE gold.fact_crashes AS
        SELECT
            crash_id,
            collision_id,
            CASE
                WHEN crash_at IS NOT NULL
                THEN strftime(CAST(crash_at AS DATE), '%Y%m%d')::INTEGER
                ELSE NULL
            END AS crash_date_key,
            borough,
            zip_code,
            latitude,
            longitude,
            persons_injured,
            persons_killed,
            pedestrians_injured,
            pedestrians_killed,
            cyclist_injured,
            cyclist_killed,
            motorist_injured,
            motorist_killed,
            vehicle_type_1,
            vehicle_type_2,
            contributing_factor_1,
            contributing_factor_2,
            is_location_missing
        FROM silver.crash_detail
        """
    )

    conn.execute("DROP TABLE IF EXISTS gold.fact_crash_persons")
    conn.execute(
        """
        CREATE TABLE gold.fact_crash_persons AS
        SELECT
            crash_person_id,
            unique_id,
            collision_id,
            CASE
                WHEN crash_date IS NOT NULL
                THEN strftime(crash_date, '%Y%m%d')::INTEGER
                ELSE NULL
            END AS crash_date_key,
            person_type,
            person_injury,
            person_sex,
            person_age,
            ped_role,
            ejection,
            emotional_status,
            bodily_injury,
            position_in_vehicle,
            safety_equipment,
            is_orphan_record,
            is_age_anomaly
        FROM silver.crash_person_detail
        """
    )

    return [
        "fact_trips",
        "fact_parking_violations",
        "fact_tif_payments",
        "fact_driver_applications",
        "fact_crashes",
        "fact_crash_persons",
    ]


def build_g3(conn) -> list[str]:
    """构建 G3 汇总表和中文语义层"""
    conn.execute("DROP TABLE IF EXISTS gold.dws_daily_trip_summary")
    conn.execute(
        """
        CREATE TABLE gold.dws_daily_trip_summary AS
        SELECT
            t.pickup_date_key AS date_key,
            d.date AS trip_date,
            count(*)::BIGINT AS trip_count,
            sum(COALESCE(t.passenger_count, 0))::BIGINT AS total_passenger_count,
            sum(COALESCE(t.distance_miles, 0))::DOUBLE AS total_distance_miles,
            sum(COALESCE(t.fare_amount, 0))::DECIMAL(18,2) AS total_fare_amount,
            avg(t.distance_miles)::DOUBLE AS avg_distance_miles,
            sum(CASE WHEN t.is_time_anomaly THEN 1 ELSE 0 END)::BIGINT AS time_anomaly_count,
            sum(CASE WHEN t.is_location_missing THEN 1 ELSE 0 END)::BIGINT AS location_missing_count,
            sum(CASE WHEN t.is_distance_outlier THEN 1 ELSE 0 END)::BIGINT AS distance_outlier_count
        FROM gold.fact_trips t
        LEFT JOIN gold.dim_date d
          ON d.date_key = t.pickup_date_key
        WHERE t.pickup_date_key IS NOT NULL
        GROUP BY t.pickup_date_key, d.date
        ORDER BY t.pickup_date_key
        """
    )

    conn.execute("DROP TABLE IF EXISTS gold.dws_zone_trip_summary")
    conn.execute(
        """
        CREATE TABLE gold.dws_zone_trip_summary AS
        SELECT
            t.pickup_location_id,
            z.borough,
            z.zone_name,
            count(*)::BIGINT AS trip_count,
            sum(COALESCE(t.distance_miles, 0))::DOUBLE AS total_distance_miles,
            sum(COALESCE(t.fare_amount, 0))::DECIMAL(18,2) AS total_fare_amount,
            avg(t.distance_miles)::DOUBLE AS avg_distance_miles,
            sum(CASE WHEN t.is_location_missing THEN 1 ELSE 0 END)::BIGINT AS location_missing_count
        FROM gold.fact_trips t
        LEFT JOIN gold.dim_taxi_zone z
          ON z.location_id = t.pickup_location_id
        WHERE t.pickup_location_id IS NOT NULL
        GROUP BY t.pickup_location_id, z.borough, z.zone_name
        ORDER BY trip_count DESC
        """
    )

    conn.execute("DROP TABLE IF EXISTS gold.dws_daily_parking_summary")
    conn.execute(
        """
        CREATE TABLE gold.dws_daily_parking_summary AS
        SELECT
            p.issue_date_key AS date_key,
            d.date AS issue_date,
            count(*)::BIGINT AS violation_count,
            count(DISTINCT p.plate_id)::BIGINT AS unique_plate_count,
            sum(COALESCE(p.standard_fine_amount, 0))::DECIMAL(18,2) AS standard_fine_total,
            count(p.standard_fine_amount)::BIGINT AS standard_fine_covered_count,
            sum(CASE WHEN p.is_duplicate_summons THEN 1 ELSE 0 END)::BIGINT AS duplicate_summons_count
        FROM gold.fact_parking_violations p
        LEFT JOIN gold.dim_date d
          ON d.date_key = p.issue_date_key
        WHERE p.issue_date_key IS NOT NULL
        GROUP BY p.issue_date_key, d.date
        ORDER BY p.issue_date_key
        """
    )

    conn.execute("DROP TABLE IF EXISTS gold.dws_daily_crash_summary")
    conn.execute(
        """
        CREATE TABLE gold.dws_daily_crash_summary AS
        SELECT
            c.crash_date_key AS date_key,
            d.date AS crash_date,
            count(*)::BIGINT AS crash_count,
            sum(COALESCE(c.persons_injured, 0))::BIGINT AS persons_injured,
            sum(COALESCE(c.persons_killed, 0))::BIGINT AS persons_killed,
            sum(COALESCE(c.pedestrians_injured, 0))::BIGINT AS pedestrians_injured,
            sum(COALESCE(c.pedestrians_killed, 0))::BIGINT AS pedestrians_killed,
            sum(COALESCE(c.cyclist_injured, 0))::BIGINT AS cyclist_injured,
            sum(COALESCE(c.cyclist_killed, 0))::BIGINT AS cyclist_killed,
            sum(COALESCE(c.motorist_injured, 0))::BIGINT AS motorist_injured,
            sum(COALESCE(c.motorist_killed, 0))::BIGINT AS motorist_killed,
            sum(CASE WHEN c.is_location_missing THEN 1 ELSE 0 END)::BIGINT AS location_missing_count
        FROM gold.fact_crashes c
        LEFT JOIN gold.dim_date d
          ON d.date_key = c.crash_date_key
        WHERE c.crash_date_key IS NOT NULL
        GROUP BY c.crash_date_key, d.date
        ORDER BY c.crash_date_key
        """
    )

    write_semantic_layer(conn)
    return [
        "dws_daily_trip_summary",
        "dws_zone_trip_summary",
        "dws_daily_parking_summary",
        "dws_daily_crash_summary",
    ]


def parse_batches(value: str) -> set[str]:
    """解析构建批次参数"""
    batches = {item.strip().upper() for item in value.split(",") if item.strip()}
    invalid = batches - {"G0", "G1", "G2", "G3"}
    if invalid:
        raise ValueError(f"当前脚本只支持 G0/G1/G2/G3，收到无效批次: {', '.join(sorted(invalid))}")
    return batches


def build_gold(db_path: Path, batches: set[str]) -> list[str]:
    """按批次构建 Gold 表"""
    conn = connect_duckdb(db_path)
    built_tables: list[str] = []
    try:
        ensure_meta_tables(conn)
        if "G0" in batches:
            built_tables.extend(build_g0(conn))
        if "G1" in batches:
            built_tables.extend(build_g1(conn))
        if "G2" in batches:
            built_tables.extend(build_g2(conn))
        if "G3" in batches:
            built_tables.extend(build_g3(conn))
        replace_gold_comments(conn, built_tables)
        write_gold_comments(conn, built_tables)
        return built_tables
    finally:
        conn.close()


def main() -> int:
    """命令行入口"""
    os.environ.setdefault("PYTHONIOENCODING", "utf-8")
    config = load_harness_config()
    parser = argparse.ArgumentParser(description="构建 DuckDB Gold G0/G1/G2/G3 表")
    parser.add_argument("--db", type=Path, default=config.duckdb_path, help="DuckDB 数据库路径")
    parser.add_argument("--batches", default="G0,G1", help="构建批次，支持 G0,G1,G2,G3")
    args = parser.parse_args()

    try:
        batches = parse_batches(args.batches)
        built_tables = build_gold(args.db, batches)
    except Exception as exc:
        print(f"[FAIL] Gold 构建失败: {exc}")
        return 1

    print("[OK] Gold 构建完成。")
    for table_name in built_tables:
        print(f"- gold.{table_name}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
