from __future__ import annotations

from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
import pyarrow.parquet as pq
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BASE_DIR = Path("D:/Program Files/Datawarehouse/" + "\u7ebd\u7ea6\u5e02\u57ce\u5e02\u4ea4\u901a")
CHUNK_ROOT = Path("D:/PycharmProjects/PySpark/data")
OUTPUT_PATH = BASE_DIR / "\u7ebd\u7ea6\u5e02\u57ce\u5e02\u4ea4\u901a_\u5168\u57df\u6570\u636e\u89c4\u8303.xlsx"
GENERATED_AT = datetime.now().strftime("%Y-%m-%d %H:%M:%S")


TABLE_ZH = {
    "yellow_tripdata": "黄色出租车行程表",
    "green_tripdata": "绿色出租车行程表",
    "fhv_tripdata": "普通网约车行程表",
    "fhvhv_tripdata": "高流量网约车行程表",
    "taxi_zone_lookup": "出租车区域维表",
    "TLC_New_Driver_Application_Status_20260605": "TLC新司机申请状态表",
    "Taxi_Improvement_Fund_(TIF)_Medallion_Payments_20260605": "出租车改善基金支付表",
    "For_Hire_Vehicles_(FHV)_-_Active_20260605": "FHV活跃车辆表",
    "For_Hire_Vehicles_(FHV)_-_Active_Drivers_20260605": "FHV活跃司机表",
    "Medallion__Vehicles_-_Authorized_20260605": "授权出租车牌照车辆表",
    "active_vehicles": "活跃车辆注册表",
    "Street_Hail_Livery_(SHL)_Drivers_-_Active_20260605": "SHL活跃司机表",
    "crash_chunks": "机动车碰撞事故事实分片表",
    "crash_person_chunks": "机动车碰撞事故人员分片表",
    "parking_violations_chunks": "停车违章罚单分片表",
}


FIELD_ZH = {
    "VendorID": "供应商编号",
    "tpep_pickup_datetime": "黄色出租车接客时间",
    "tpep_dropoff_datetime": "黄色出租车送客时间",
    "lpep_pickup_datetime": "绿色出租车接客时间",
    "lpep_dropoff_datetime": "绿色出租车送客时间",
    "pickup_datetime": "接客时间",
    "dropOff_datetime": "送客时间",
    "dropoff_datetime": "送客时间",
    "request_datetime": "乘客请求时间",
    "on_scene_datetime": "车辆到达现场时间",
    "PULocationID": "上车区域编号",
    "DOLocationID": "下车区域编号",
    "PUlocationID": "上车区域编号",
    "DOlocationID": "下车区域编号",
    "passenger_count": "乘客人数",
    "trip_distance": "行程距离",
    "trip_miles": "行程英里数",
    "trip_time": "行程时长秒数",
    "fare_amount": "车费金额",
    "base_passenger_fare": "基础乘客费用",
    "total_amount": "总金额",
    "tip_amount": "小费金额",
    "tips": "小费金额",
    "tolls_amount": "过路费金额",
    "tolls": "过路费金额",
    "driver_pay": "司机收入",
    "payment_type": "支付方式",
    "congestion_surcharge": "拥堵附加费",
    "cbd_congestion_fee": "中央商务区拥堵费",
    "Airport_fee": "机场费",
    "airport_fee": "机场费",
    "hvfhs_license_num": "高流量网约车牌照号",
    "dispatching_base_num": "派车基地编号",
    "originating_base_num": "原始基地编号",
    "Affiliated_base_number": "关联基地编号",
    "SR_Flag": "共享乘车标记",
    "shared_request_flag": "请求共享标记",
    "shared_match_flag": "成功拼车标记",
    "access_a_ride_flag": "无障碍接驳标记",
    "wav_request_flag": "无障碍车辆请求标记",
    "wav_match_flag": "无障碍车辆匹配标记",
    "LocationID": "出租车区域编号",
    "Borough": "行政区",
    "Zone": "出租车区域",
    "service_zone": "服务区域",
    "Active": "是否活跃",
    "Vehicle License Number": "车辆牌照号",
    "License Number": "牌照号",
    "License Type": "牌照类型",
    "Expiration Date": "到期日期",
    "DMV License Plate Number": "DMV车牌号",
    "DMV Plate Number": "DMV车牌号",
    "Vehicle VIN Number": "车辆VIN码",
    "VIN": "车辆VIN码",
    "Vehicle Year": "车辆年份",
    "Vehicle Make": "车辆品牌",
    "Vehicle Model": "车辆型号",
    "Fuel Type": "燃料类型",
    "Wheelchair Accessible": "是否无障碍车辆",
    "Wheelchair Accessible Trained": "是否完成无障碍培训",
    "WAV": "无障碍车辆标记",
    "Base Number": "基地编号",
    "Base Name": "基地名称",
    "Base Type": "基地类型",
    "Base Address": "基地地址",
    "Owner Name": "车主名称",
    "TLC Vehicle License Status": "TLC车辆牌照状态",
    "Affiliated Base/ Agent Number": "关联基地或代理编号",
    "Current Status": "当前状态",
    "Medallion Type": "出租车牌照类型",
    "Agent Number": "代理编号",
    "Agent Name": "代理名称",
    "App No": "申请编号",
    "Type": "类型",
    "App Date": "申请日期",
    "Status": "状态",
    "Last Updated": "最后更新时间",
    "Last Date Updated": "最后更新日期",
    "Last Time Updated": "最后更新时间",
    "Payment Date": "支付日期",
    "Total Payment Amount": "总支付金额",
    "Hackup Payment Amount": "上牌改装支付金额",
    "Operational Payment Amount": "运营支付金额",
    "crash_date": "事故日期",
    "crash_time": "事故时间",
    "collision_id": "事故编号",
    "borough": "行政区",
    "zip_code": "邮编",
    "latitude": "纬度",
    "longitude": "经度",
    "location": "地理位置",
    "on_street_name": "所在街道",
    "cross_street_name": "交叉街道",
    "off_street_name": "相邻街道",
    "number_of_persons_injured": "受伤人数",
    "number_of_persons_killed": "死亡人数",
    "number_of_pedestrians_injured": "行人受伤人数",
    "number_of_pedestrians_killed": "行人死亡人数",
    "number_of_cyclist_injured": "骑行者受伤人数",
    "number_of_cyclist_killed": "骑行者死亡人数",
    "number_of_motorist_injured": "机动车乘员受伤人数",
    "number_of_motorist_killed": "机动车乘员死亡人数",
    "contributing_factor_vehicle_1": "车辆1碰撞因素",
    "contributing_factor_vehicle_2": "车辆2碰撞因素",
    "vehicle_type_code1": "车辆1类型",
    "vehicle_type_code2": "车辆2类型",
    "unique_id": "人员记录编号",
    "person_id": "人员编号",
    "person_type": "人员类型",
    "person_injury": "人员伤害程度",
    "vehicle_id": "事故车辆编号",
    "person_age": "人员年龄",
    "person_sex": "人员性别",
    "summons_number": "罚单编号",
    "plate_id": "车牌号",
    "registration_state": "注册州",
    "plate_type": "车牌类型",
    "issue_date": "开票日期",
    "violation_code": "违章代码",
    "violation_description": "违章描述",
    "violation_time": "违章时间",
    "street_name": "街道名称",
    "violation_county": "违章县区",
    "fiscal_year": "财年",
}


def table_key(name: str) -> str:
    for prefix in ["yellow_tripdata", "green_tripdata", "fhv_tripdata", "fhvhv_tripdata"]:
        if name.startswith(prefix):
            return prefix
    return name


def table_zh(name: str) -> str:
    return TABLE_ZH.get(table_key(name), TABLE_ZH.get(name, f"待补充中文表名：{name}"))


def field_zh(name: str) -> str:
    return FIELD_ZH.get(name, f"待补充：{name}")


def type_label(raw_type: str) -> str:
    lower = raw_type.lower()
    if "timestamp" in lower or "date" in lower:
        return "日期时间"
    if any(token in lower for token in ["int", "double", "float", "decimal"]):
        return "数值"
    if "struct" in lower:
        return "结构对象"
    return "文本"


def domain_for(name: str, folder: str) -> str:
    text = f"{name} {folder}".lower()
    if any(token in text for token in ["tripdata"]):
        return "出行域"
    if "taxi_zone_lookup" in text or "taxi_zones" in text:
        return "空间地理域"
    if any(token in text for token in ["crash", "collision", "事故"]):
        return "安全域"
    if any(token in text for token in ["parking", "violation", "tif", "authorized", "application", "罚单", "授权"]):
        return "监管合规域"
    if any(token in text for token in ["drivers", "driver", "司机"]):
        return "供给域"
    if any(token in text for token in ["vehicle", "vehicles", "medallion", "active_vehicles"]):
        return "资产域"
    return "待归类"


def role_for(name: str, folder: str) -> str:
    text = f"{name} {folder}".lower()
    if "tripdata" in text:
        return "行程事实表"
    if "taxi_zone_lookup" in text:
        return "空间维表"
    if "crash_person" in text:
        return "事故人员明细表"
    if "crash" in text or "collision" in text:
        return "事故事实表"
    if "parking" in text:
        return "罚单事实表"
    if "tif" in text:
        return "支付事实表"
    if "application" in text:
        return "申请状态事实表"
    if "drivers" in text or "司机" in text:
        return "司机快照表"
    if "vehicle" in text or "medallion" in text:
        return "车辆/牌照维表"
    return "参考表"


def primary_key_hint(name: str, columns: list[str]) -> str:
    for key in ["unique_id", "collision_id", "summons_number", "LocationID", "App No", "Vehicle License Number", "License Number"]:
        if key in columns:
            return key
    if "tripdata" in name:
        return "无自然主键，建议生成 trip_id"
    return "需确认"


def relation_key(columns: list[str]) -> str:
    keys = []
    for column in columns:
        low = column.lower()
        if any(token in low for token in ["locationid", "collision_id", "license", "plate", "vin", "base", "summons", "app no"]):
            keys.append(column)
    return " / ".join(keys[:8]) if keys else "需在标准层补充或生成"


def field_level(field: str, primary_key: str) -> str:
    low = field.lower()
    if field == primary_key:
        return "主键"
    if any(token in low for token in ["locationid", "collision_id", "license", "plate", "vin", "base", "summons"]):
        return "关联键"
    if any(token in low for token in ["date", "time", "datetime"]):
        return "时间字段"
    if any(token in low for token in ["amount", "fare", "fee", "toll", "tax", "pay", "surcharge"]):
        return "金额字段"
    if any(token in low for token in ["latitude", "longitude", "borough", "street", "zone", "zip"]):
        return "空间字段"
    if any(token in low for token in ["status", "flag", "type", "code"]):
        return "状态/枚举字段"
    return "分析字段"


def empty_stats(row_count: int) -> dict[str, Any]:
    return {"缺失数": "", "缺失率": "", "非空数": row_count, "唯一值数": "", "状态": "待采样"}


def parquet_missing_stats(path: Path) -> dict[str, dict[str, Any]]:
    parquet_file = pq.ParquetFile(path)
    row_count = parquet_file.metadata.num_rows
    result: dict[str, dict[str, Any]] = {}
    names = [field.name for field in parquet_file.schema_arrow]
    for index, name in enumerate(names):
        null_count = 0
        known = True
        for group_index in range(parquet_file.metadata.num_row_groups):
            stats = parquet_file.metadata.row_group(group_index).column(index).statistics
            if stats is None or stats.null_count is None:
                known = False
                break
            null_count += stats.null_count
        if known:
            rate = null_count / row_count if row_count else 0
            result[name] = {
                "缺失数": null_count,
                "缺失率": rate,
                "非空数": row_count - null_count,
                "唯一值数": "",
                "状态": status_for_missing(rate),
            }
        else:
            result[name] = empty_stats(row_count)
    return result


def csv_missing_stats(path: Path) -> dict[str, dict[str, Any]]:
    df = pd.read_csv(path, encoding="utf-8-sig", dtype=str)
    result = {}
    for column in df.columns:
        missing = int(df[column].isna().sum())
        total = len(df)
        rate = missing / total if total else 0
        result[column] = {
            "缺失数": missing,
            "缺失率": rate,
            "非空数": total - missing,
            "唯一值数": int(df[column].nunique(dropna=True)),
            "状态": status_for_missing(rate),
        }
    return result


def status_for_missing(rate: float) -> str:
    if rate == 0:
        return "良好"
    if rate < 0.3:
        return "轻微缺失"
    if rate < 0.8:
        return "缺失偏高"
    return "严重缺失"


def collect_main() -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    summaries = []
    fields = []
    missing_rows = []
    quality = []
    for path in sorted(BASE_DIR.rglob("*.parquet")):
        parquet_file = pq.ParquetFile(path)
        columns = [(field.name, str(field.type)) for field in parquet_file.schema_arrow]
        column_names = [name for name, _ in columns]
        row_count = parquet_file.metadata.num_rows
        primary_key = primary_key_hint(path.name, column_names)
        summary = {
            "英文表名": path.stem,
            "中文表名": table_zh(path.stem),
            "目录": path.parent.name,
            "数据域": domain_for(path.name, path.parent.name),
            "文件格式": "parquet",
            "行数": row_count,
            "列数": len(columns),
            "大小MB": round(path.stat().st_size / 1024 / 1024, 2),
            "数据角色": role_for(path.name, path.parent.name),
            "建议主键": primary_key,
            "主要关联键": relation_key(column_names),
            "完整性状态": "结构完整",
            "备注": "源表文件，已纳入本次规范",
        }
        summaries.append(summary)
        stats = parquet_missing_stats(path)
        for index, (name, raw_type) in enumerate(columns, 1):
            item = stats[name]
            field_row = {
                "数据集": path.stem,
                "序号": index,
                "字段名": name,
                "中文字段名": field_zh(name),
                "标准类型": type_label(raw_type),
                "原始类型": raw_type,
                "字段层级": field_level(name, primary_key),
                "业务说明": field_zh(name),
                "治理备注": "",
            }
            fields.append(field_row)
            missing_rows.append(build_missing_row(summary, field_row, item))
    for path in sorted(BASE_DIR.rglob("*.csv")):
        df_head = pd.read_csv(path, nrows=0, encoding="utf-8-sig")
        columns = list(df_head.columns)
        with path.open("r", encoding="utf-8-sig", errors="replace") as file:
            row_count = max(sum(1 for _ in file) - 1, 0)
        primary_key = primary_key_hint(path.name, columns)
        summary = {
            "英文表名": path.stem,
            "中文表名": table_zh(path.stem),
            "目录": path.parent.name,
            "数据域": domain_for(path.name, path.parent.name),
            "文件格式": "csv",
            "行数": row_count,
            "列数": len(columns),
            "大小MB": round(path.stat().st_size / 1024 / 1024, 2),
            "数据角色": role_for(path.name, path.parent.name),
            "建议主键": primary_key,
            "主要关联键": relation_key(columns),
            "完整性状态": "结构完整",
            "备注": "源表文件，已纳入本次规范",
        }
        summaries.append(summary)
        stats = csv_missing_stats(path)
        for index, name in enumerate(columns, 1):
            item = stats[name]
            note = ""
            if item["缺失率"] >= 0.8:
                note = "高缺失字段，需标注适用范围"
            elif item["缺失率"] >= 0.3:
                note = "缺失率偏高，分析时需谨慎"
            field_row = {
                "数据集": path.stem,
                "序号": index,
                "字段名": name,
                "中文字段名": field_zh(name),
                "标准类型": "文本/待推断",
                "原始类型": "csv文本",
                "字段层级": field_level(name, primary_key),
                "业务说明": field_zh(name),
                "治理备注": note,
            }
            fields.append(field_row)
            missing_rows.append(build_missing_row(summary, field_row, item))
            if note:
                quality.append(
                    {
                        "数据集": path.stem,
                        "严重度": "高" if item["缺失率"] >= 0.8 else "中",
                        "问题类型": "字段缺失",
                        "问题描述": f"{name} 缺失率约 {item['缺失率']:.1%}",
                        "影响字段": name,
                        "建议处理": note,
                    }
                )
    return summaries, fields, missing_rows, quality


def build_missing_row(summary: dict[str, Any], field_row: dict[str, Any], stats: dict[str, Any]) -> dict[str, Any]:
    rate = stats["缺失率"]
    rate_text = "" if rate == "" else f"{rate:.2%}"
    return {
        "英文表名": summary["英文表名"],
        "中文表名": summary["中文表名"],
        "数据域": summary["数据域"],
        "字段名": field_row["字段名"],
        "中文字段名": field_row["中文字段名"],
        "标准类型": field_row["标准类型"],
        "行数": summary["行数"],
        "缺失数": stats["缺失数"],
        "缺失率": rate_text,
        "非空数": stats["非空数"],
        "唯一值数": stats["唯一值数"],
        "状态": stats["状态"],
    }


def collect_chunks() -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    summaries = []
    fields = []
    missing_rows = []
    quality = []
    for folder in ["crash_chunks", "crash_person_chunks", "parking_violations_chunks"]:
        root = CHUNK_ROOT / folder
        files = sorted(root.rglob("*.parquet"))
        if not files:
            continue
        union_columns = []
        type_variants: dict[str, set[str]] = defaultdict(set)
        row_count = 0
        schema_orders = set()
        for path in files:
            parquet_file = pq.ParquetFile(path)
            row_count += parquet_file.metadata.num_rows
            schema_orders.add(tuple(field.name for field in parquet_file.schema_arrow))
            for field in parquet_file.schema_arrow:
                if field.name not in union_columns:
                    union_columns.append(field.name)
                type_variants[field.name].add(str(field.type))
        primary_key = primary_key_hint(folder, union_columns)
        summary = {
            "英文表名": folder,
            "中文表名": table_zh(folder),
            "目录": str(root),
            "数据域": domain_for(folder, folder),
            "文件格式": "parquet分片",
            "行数": row_count,
            "列数": len(union_columns),
            "大小MB": round(sum(path.stat().st_size for path in files) / 1024 / 1024, 2),
            "数据角色": role_for(folder, folder),
            "建议主键": primary_key,
            "主要关联键": relation_key(union_columns),
            "完整性状态": "采集中，不能视为最终全量",
            "备注": f"当前 {len(files)} 个分片，字段顺序变体 {len(schema_orders)} 个",
        }
        summaries.append(summary)
        for index, name in enumerate(union_columns, 1):
            raw_type = " / ".join(sorted(type_variants[name]))
            note = "采集完成后复核"
            if len(type_variants[name]) > 1:
                note = "字段存在类型漂移，需统一"
            field_row = {
                "数据集": folder,
                "序号": index,
                "字段名": name,
                "中文字段名": field_zh(name),
                "标准类型": type_label(raw_type),
                "原始类型": raw_type,
                "字段层级": field_level(name, primary_key),
                "业务说明": field_zh(name),
                "治理备注": note,
            }
            fields.append(field_row)
            missing_rows.append(
                {
                    "英文表名": folder,
                    "中文表名": table_zh(folder),
                    "数据域": summary["数据域"],
                    "字段名": name,
                    "中文字段名": field_zh(name),
                    "标准类型": type_label(raw_type),
                    "行数": row_count,
                    "缺失数": "",
                    "缺失率": "",
                    "非空数": "",
                    "唯一值数": "",
                    "状态": "采集中，待下载完成后统计",
                }
            )
        if len(schema_orders) > 1:
            quality.append(
                {
                    "数据集": folder,
                    "严重度": "高",
                    "问题类型": "schema漂移",
                    "问题描述": f"分片字段顺序或字段集合存在 {len(schema_orders)} 个变体",
                    "影响字段": "全表",
                    "建议处理": "入湖前按 union schema 固定列顺序，缺失列补空",
                }
            )
    return summaries, fields, missing_rows, quality


def append_plain_sheet(workbook: Workbook, title: str, rows: list[dict[str, Any]] | list[list[Any]]) -> None:
    sheet = workbook.create_sheet(title)
    if not rows:
        sheet.append(["无数据"])
        return
    if isinstance(rows[0], dict):
        headers = list(rows[0].keys())
        sheet.append(headers)
        for row in rows:
            sheet.append([row.get(header, "") for header in headers])
    else:
        for row in rows:
            sheet.append(row)
    style_sheet(sheet)


def append_dictionary_sheet(workbook: Workbook, summaries: list[dict[str, Any]], fields: list[dict[str, Any]]) -> None:
    sheet = workbook.create_sheet("数据字典")
    headers = ["序号", "字段名", "中文字段名", "标准类型", "原始类型", "字段层级", "业务说明", "治理备注"]
    fields_by_table: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for field in fields:
        fields_by_table[field["数据集"]].append(field)
    current_row = 1
    for summary in summaries:
        table_fields = fields_by_table.get(summary["英文表名"], [])
        if not table_fields:
            continue
        title = f"{summary['英文表名']}({summary['中文表名']})"
        sheet.cell(current_row, 1, title)
        sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
        title_cell = sheet.cell(current_row, 1)
        title_cell.font = Font(bold=True, size=13, color="FFFFFF")
        title_cell.fill = PatternFill("solid", fgColor="1F4E78")
        title_cell.alignment = Alignment(vertical="center")
        current_row += 1
        description = (
            f"表说明：数据域={summary['数据域']}；数据角色={summary['数据角色']}；"
            f"行数={summary['行数']}；列数={summary['列数']}；主键/候选键={summary['建议主键']}；"
            f"主要关联键={summary['主要关联键']}；完整性状态={summary['完整性状态']}。"
        )
        sheet.cell(current_row, 1, description)
        sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
        sheet.cell(current_row, 1).alignment = Alignment(wrap_text=True, vertical="top")
        current_row += 1
        for index, header in enumerate(headers, 1):
            cell = sheet.cell(current_row, index, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="5B9BD5")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        current_row += 1
        for field in table_fields:
            values = [field.get(header, "") for header in headers]
            for index, value in enumerate(values, 1):
                sheet.cell(current_row, index, value)
                sheet.cell(current_row, index).alignment = Alignment(vertical="top", wrap_text=True)
            current_row += 1
        current_row += 1
    sheet.freeze_panes = "A1"
    apply_widths(sheet)


def style_sheet(sheet: Any) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A2"
    apply_widths(sheet)


def apply_widths(sheet: Any) -> None:
    for index, column in enumerate(sheet.columns, 1):
        width = min(max(len(str(cell.value)) if cell.value is not None else 0 for cell in column) + 2, 52)
        sheet.column_dimensions[get_column_letter(index)].width = max(width, 12)


def build_workbook() -> None:
    main_summaries, main_fields, main_missing, main_quality = collect_main()
    chunk_summaries, chunk_fields, chunk_missing, chunk_quality = collect_chunks()
    summaries = main_summaries + chunk_summaries
    fields = main_fields + chunk_fields
    missing_rows = main_missing + chunk_missing
    quality = main_quality + chunk_quality

    workbook = Workbook()
    workbook.remove(workbook.active)
    append_plain_sheet(workbook, "数据总览", summaries)
    append_plain_sheet(
        workbook,
        "六大数据域",
        [
            {"数据域": "出行域", "中文说明": "描述城市交通出行行为，包括上车、下车、费用、距离、时长", "代表数据": "Yellow/Green/FHV/HVFHV 行程", "建模归属": "fact_trips"},
            {"数据域": "资产域", "中文说明": "描述车辆、牌照、VIN、燃料、无障碍能力等交通资产", "代表数据": "active_vehicles、FHV车辆、Medallion车辆", "建模归属": "dim_vehicle"},
            {"数据域": "供给域", "中文说明": "描述司机、基地、申请人等可服务资源", "代表数据": "FHV司机、SHL司机、新司机申请、Base字段", "建模归属": "dim_driver / dim_base / fact_driver_applications"},
            {"数据域": "安全域", "中文说明": "描述交通事故、伤亡、人员、碰撞因素", "代表数据": "crash_chunks、crash_person_chunks", "建模归属": "fact_crashes / fact_crash_persons"},
            {"数据域": "监管合规域", "中文说明": "描述罚单、牌照状态、车辆授权、补贴支付、申请审批", "代表数据": "parking violations、TIF、授权车辆、申请状态", "建模归属": "fact_parking_violations / fact_tif_payments / dim_license_status"},
            {"数据域": "空间地理域", "中文说明": "描述区域、行政区、经纬度、街道、邮编等空间信息", "代表数据": "taxi_zone_lookup、taxi_zones.shp、经纬度、街道", "建模归属": "dim_taxi_zone"},
        ],
    )
    append_dictionary_sheet(workbook, summaries, fields)
    append_plain_sheet(workbook, "缺失率矩阵", missing_rows)
    append_plain_sheet(
        workbook,
        "表关系模型",
        [
            ["模型对象", "中文注释", "性质"],
            ["Bronze 原始层", "原始文件保留层，不改字段、不改格式，用于追溯来源", "规划层"],
            ["Silver 标准层", "清洗标准化层，统一字段名、类型、时间、空间、状态码", "规划层"],
            ["Gold 主题星型模型", "面向分析和问数的主题模型，事实表围绕公共维表组织", "规划模型"],
            ["fact_trips", "出行事实表，记录出租车、网约车等每一次行程", "规划表"],
            ["fact_parking_violations", "停车罚单事实表，记录每一张违章罚单", "规划表"],
            ["fact_tif_payments", "出租车改善基金支付事实表，记录补贴或支付行为", "规划表"],
            ["fact_driver_applications", "司机申请事实表，记录新司机申请及审批状态", "规划表"],
            ["dim_vehicle", "车辆维表，描述车辆、VIN、车牌、燃料、WAV等资产属性", "规划表"],
            ["dim_driver", "司机维表，描述司机牌照、司机类型、有效期等供给属性", "规划表"],
            ["dim_base", "基地维表，描述FHV基地、基地类型、联系方式和地址", "规划表"],
            ["dim_taxi_zone", "出租车区域维表，描述LocationID、行政区、服务区域", "规划表"],
            ["dim_date", "日期维表，统一支持日、周、月、季度、财年等时间分析", "规划表"],
            ["dim_violation_code", "违章代码维表，解释罚单违章类型和罚款金额", "规划表"],
            ["dim_license_status", "牌照状态维表，统一解释车辆、司机、申请状态", "规划表"],
            ["事故 ER 子模型", "安全域保留事件与人员明细的一对多关系", "规划模型"],
            ["fact_crashes", "事故事实表，记录每一次碰撞事故的位置、时间、伤亡和原因", "规划表"],
            ["fact_crash_persons", "事故人员明细表，记录事故涉及人员、伤害程度和人员属性", "规划表"],
            ["中文语义层", "规划层/元数据层，用于未来支撑中文用户问数和Agent理解，不是当前源数据表", "规划层"],
        ],
    )
    append_plain_sheet(
        workbook,
        "入湖标准字段映射",
        [
            {"原始字段": "tpep_pickup_datetime/lpep_pickup_datetime/pickup_datetime", "标准字段": "pickup_at", "中文注释": "统一接客时间"},
            {"原始字段": "tpep_dropoff_datetime/lpep_dropoff_datetime/dropoff_datetime/dropOff_datetime", "标准字段": "dropoff_at", "中文注释": "统一送客时间"},
            {"原始字段": "PULocationID/PUlocationID", "标准字段": "pickup_location_id", "中文注释": "统一上车区域编号"},
            {"原始字段": "DOLocationID/DOlocationID", "标准字段": "dropoff_location_id", "中文注释": "统一下车区域编号"},
            {"原始字段": "License Number/Vehicle License Number", "标准字段": "license_no", "中文注释": "统一牌照号"},
            {"原始字段": "DMV License Plate Number/DMV Plate Number/plate_id", "标准字段": "plate_no", "中文注释": "统一车牌号"},
            {"原始字段": "VIN/Vehicle VIN Number", "标准字段": "vehicle_vin", "中文注释": "统一车辆VIN"},
            {"原始字段": "Base Number/dispatching_base_num/Affiliated Base/ Agent Number", "标准字段": "base_no", "中文注释": "统一基地编号"},
            {"原始字段": "collision_id", "标准字段": "collision_id", "中文注释": "事故主关联键"},
            {"原始字段": "summons_number", "标准字段": "summons_no", "中文注释": "罚单主键"},
        ],
    )
    append_plain_sheet(
        workbook,
        "中文语义层",
        [
            {"对象": "中文语义层", "中文注释": "规划层/元数据层，用于未来支撑中文用户问数和Agent理解，不是当前源数据表", "性质": "规划"},
            {"对象": "semantic_tables", "中文注释": "中文表说明，解释每张表能回答什么问题", "性质": "规划元数据"},
            {"对象": "semantic_fields", "中文注释": "中文字段说明，解释字段含义、单位、适用范围", "性质": "规划元数据"},
            {"对象": "semantic_metrics", "中文注释": "中文指标口径，如行程量、事故率、罚单量、WAV覆盖率", "性质": "规划元数据"},
            {"对象": "semantic_relationships", "中文注释": "中文关系说明，标注强关联、弱关联和不可关联", "性质": "规划元数据"},
            {"对象": "semantic_question_templates", "中文注释": "中文问数模板，约束Agent如何安全生成查询", "性质": "规划元数据"},
        ],
    )
    append_plain_sheet(workbook, "数据质量问题", quality)
    workbook.save(OUTPUT_PATH)


if __name__ == "__main__":
    build_workbook()
