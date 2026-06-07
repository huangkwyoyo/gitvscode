"""生成 Silver 层数据字典 xlsx"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
p0_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
p1_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
p2_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
cell_align = Alignment(vertical='center', wrap_text=True)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

wb = openpyxl.Workbook()

# ===== Sheet 1: 总览 =====
ws0 = wb.active
ws0.title = 'Silver表清单'
overview_headers = ['批次', '英文表名', '中文表名', '数据域', '数据角色', '字段数', '预计行数', '主键/候选键', '来源Bronze表']
for col, h in enumerate(overview_headers, 1):
    c = ws0.cell(row=1, column=col, value=h)
    c.font = header_font; c.fill = header_fill; c.alignment = header_align; c.border = thin_border

overview = [
    ['P0', 'silver.dim_date', '日期维表', '通用', '维表', 10, '~90', 'date_key', '从 trip_detail 日期范围生成'],
    ['P0', 'silver.taxi_zone', '出租车区域标准维表', '空间地理域', '维表', 5, '265', 'location_id', 'bronze.taxi_zone_lookup'],
    ['P0', 'silver.trip_detail', '行程明细标准表', '出行域', '事实表', 42, '8,032万', 'trip_id（代理键）', '四类TLC行程表 UNION ALL'],
    ['P1', 'silver.vehicle_detail', '车辆明细标准表', '资产域', '维表', 25, '~12万', 'vehicle_id（代理键）', 'active_vehicles + fhv_active_vehicles + medallion_authorized_vehicles'],
    ['P1', 'silver.driver_detail', '司机明细标准表', '供给域', '维表', 11, '~36万', 'license_number + driver_type', 'fhv_active_drivers + shl_active_drivers'],
    ['P1', 'silver.base_detail', '基地月度明细标准表', '供给域', '事实表', 12, '~5.9万', 'composite_key', 'bronze.fhv_base_aggregate_report'],
    ['P1', 'silver.driver_application_detail', '司机申请明细标准表', '监管合规域', '事实表', 14, '4,076', 'app_no', 'bronze.new_driver_applications'],
    ['P2', 'silver.parking_violation_detail', '停车罚单明细标准表', '监管合规域', '事实表', 36, '958万', 'violation_id（代理键）', 'bronze.parking_violations_all'],
    ['P2', 'silver.tif_payment_detail', 'TIF支付明细标准表', '监管合规域', '事实表', 11, '~4.8万', 'composite_key', 'bronze.tif_medallion_payments'],
    ['P2', 'silver.crash_detail', '事故明细标准表', '安全域', '事实表', 22, '166万', 'collision_id', 'bronze.crash_merged'],
    ['P2', 'silver.crash_person_detail', '事故人员明细标准表', '安全域', '事实表', 20, '533万', 'unique_id', 'bronze.crash_person_all'],
]

fill_map = {'P0': p0_fill, 'P1': p1_fill, 'P2': p2_fill}
for r, row in enumerate(overview, 2):
    batch = row[0]
    for c, val in enumerate(row, 1):
        cell = ws0.cell(row=r, column=c, value=val)
        cell.alignment = cell_align; cell.border = thin_border
        cell.fill = fill_map.get(batch, PatternFill())

ws0.column_dimensions['A'].width = 8
ws0.column_dimensions['B'].width = 34
ws0.column_dimensions['C'].width = 24
ws0.column_dimensions['D'].width = 14
ws0.column_dimensions['E'].width = 10
ws0.column_dimensions['F'].width = 10
ws0.column_dimensions['G'].width = 14
ws0.column_dimensions['H'].width = 28
ws0.column_dimensions['I'].width = 48

# ===== 辅助函数 =====
FIELD_HEADERS = ['英文表名', '中文表名', '英文字段名', '中文字段名', '数据类型', '字段层级', '业务含义', '治理备注']

def make_sheet(wb, sheet_name, table_en, table_zh, fields):
    """创建一个表的字段字典 sheet"""
    ws = wb.create_sheet(title=sheet_name)
    for col, h in enumerate(FIELD_HEADERS, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font; c.fill = header_fill; c.alignment = header_align; c.border = thin_border

    for r, f in enumerate(fields, 2):
        row_data = [table_en, table_zh] + list(f)
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = cell_align; cell.border = thin_border

    ws.column_dimensions['A'].width = 34
    ws.column_dimensions['B'].width = 24
    ws.column_dimensions['C'].width = 28
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 48
    ws.column_dimensions['H'].width = 36
    ws.freeze_panes = 'A2'

# ===== 各表字段数据 =====

make_sheet(wb, 'dim_date', 'silver.dim_date', '日期维表', [
    ('date_key', '日期键', 'INTEGER', '主键', 'YYYYMMDD 格式整数，如 20260115', '不允许为空，必须唯一'),
    ('date', '日期', 'DATE', '维度属性', '标准日期类型', ''),
    ('year', '年', 'INTEGER', '维度属性', '如 2026', ''),
    ('quarter', '季度', 'INTEGER', '维度属性', '1-4', ''),
    ('month', '月', 'INTEGER', '维度属性', '1-12', ''),
    ('week', '周', 'INTEGER', '维度属性', 'ISO 周号 1-53', ''),
    ('day_of_week', '星期几', 'INTEGER', '维度属性', '1=周一，7=周日', ''),
    ('day_of_week_name', '星期名称', 'VARCHAR', '维度属性', 'Monday-Sunday 英文名', ''),
    ('is_weekend', '是否周末', 'BOOLEAN', '维度属性', '周六/周日为 TRUE', ''),
    ('fiscal_year', 'NYC财年', 'INTEGER', '维度属性', 'NYC 财年：7月1日-6月30日。2026年1-3月属于 FY2026', ''),
])

make_sheet(wb, 'taxi_zone', 'silver.taxi_zone', '出租车区域标准维表', [
    ('location_id', '出租车区域编号', 'INTEGER', '主键', '区域唯一编号 1-265，来源于 LocationID', '不允许为空，必须唯一'),
    ('borough', '行政区', 'VARCHAR', '维度属性', 'Manhattan / Queens / Brooklyn / Bronx / Staten Island / EWR / Unknown', '保留英文原值'),
    ('zone_name', '区域名称', 'VARCHAR', '维度属性', '261 个唯一区域名称，如 Upper East Side', '保留英文原值'),
    ('service_zone', '服务区域', 'VARCHAR', '维度属性', 'Yellow Zone / Boro Zone / Airports / EWR', '保留英文原值'),
    ('is_unknown_zone', '是否未知区域', 'BOOLEAN', '派生字段', 'borough = Unknown 时为 TRUE', '用于过滤无效空间分析'),
])

# trip_detail 字段（42个）
trip_fields = [
    ('trip_id', '行程代理主键', 'VARCHAR', '主键', '格式 {trip_source}_{row_number}，如 yellow_0000001', 'Silver 层生成，不来自源表'),
    ('trip_source', '行程来源类型', 'VARCHAR', '退化维度', 'yellow / green / fhv / fhvhv', '不允许为空'),
    # 时间字段
    ('pickup_at', '接客时间', 'TIMESTAMP', '时间字段', '统一四个源表的上车时间字段名', '不允许为空'),
    ('dropoff_at', '送客时间', 'TIMESTAMP', '时间字段', '统一下车时间字段名', '允许为空，为空时标记时间异常'),
    ('pickup_date', '接客日期', 'DATE', '时间字段', '从 pickup_at 派生，关联 dim_date.date', ''),
    # 空间字段
    ('pickup_location_id', '上车区域编号', 'INTEGER', '空间字段', '统一 PULocationID / PUlocationID，关联 taxi_zone', 'FHV 类缺失率 87%，标记 is_location_missing'),
    ('dropoff_location_id', '下车区域编号', 'INTEGER', '空间字段', '统一 DOLocationID / DOlocationID', ''),
    # 维度外键
    ('base_no', '派车基地编号', 'VARCHAR', '维度外键', 'FHV/HVFHV 可用，关联 vehicle_detail.base_number', 'Yellow/Green 为 NULL'),
    # 度量字段
    ('passenger_count', '乘客人数', 'BIGINT', '度量', 'Yellow/Green/HVFHV 可用，FHV 无此字段', 'Yellow 缺失率 24-30%'),
    ('distance_miles', '行程距离（英里）', 'DOUBLE', '度量', 'Yellow/Green 用 trip_distance，HVFHV 用 trip_miles', '极端值达 328,522，需标记异常'),
    ('payment_type', '支付方式', 'BIGINT', '度量', 'Yellow/Green 可用。0=Flex,1=信用卡,2=现金,3=免费,4=争议,5=未知', 'FHV/HVFHV 为 NULL'),
    ('rate_code_id', '费率代码', 'BIGINT', '度量', 'Yellow/Green 可用。1=标准,2=JFK,3=Newark,4=Nassau,5=议价,6=拼车,99=未知', ''),
    ('trip_type', '行程类型', 'BIGINT', '度量', 'Green 独有。1=路边拦车,2=调度', '其他来源为 NULL'),
    # 金额字段（统一 DECIMAL）
    ('fare_amount', '基础车费', 'DECIMAL(12,2)', '金额', 'Yellow/Green 用 fare_amount，HVFHV 用 base_passenger_fare', '出现负值时标记异常'),
    ('total_amount', '总费用', 'DECIMAL(12,2)', '金额', 'Yellow/Green 可用', '不含现金小费（Yellow）'),
    ('extra', '附加费', 'DECIMAL(12,2)', '金额', 'Yellow/Green 可用，夜间/高峰期附加费', ''),
    ('mta_tax', 'MTA税', 'DECIMAL(12,2)', '金额', 'Yellow/Green 可用', ''),
    ('tip_amount', '小费', 'DECIMAL(12,2)', '金额', 'Yellow/Green 用 tip_amount，HVFHV 用 tips', '信用卡自动小费'),
    ('tolls_amount', '通行费', 'DECIMAL(12,2)', '金额', 'Yellow/Green 用 tolls_amount，HVFHV 用 tolls', ''),
    ('improvement_surcharge', '改善附加费', 'DECIMAL(12,2)', '金额', 'Yellow/Green 可用', 'TLC 改善基金附加费'),
    ('congestion_surcharge', '拥堵附加费', 'DECIMAL(12,2)', '金额', 'Yellow/Green/HVFHV 共有', ''),
    ('airport_fee', '机场费', 'DECIMAL(12,2)', '金额', 'Yellow/Green/HVFHV 共有', 'LGA/JFK 机场费用'),
    ('cbd_congestion_fee', 'CBD拥堵费', 'DECIMAL(12,2)', '金额', 'Yellow/Green/HVFHV 共有', '2025年1月5日起新增'),
    ('sales_tax', '销售税', 'DECIMAL(12,2)', '金额', 'HVFHV 独有', '纽约州销售税'),
    ('bcf', '黑车基金费', 'DECIMAL(12,2)', '金额', 'HVFHV 独有', 'Black Car Fund 附加费'),
    ('driver_pay', '司机净收入', 'DECIMAL(12,2)', '金额', 'HVFHV 独有', '司机实际所得，本表独有字段'),
    ('ehail_fee', '电子预约费', 'DECIMAL(12,2)', '金额', 'Green 独有', '100% 为空，已实质废弃'),
    # HVFHV 独有标志位和时间
    ('request_datetime', '乘客请求时间', 'TIMESTAMP', '时间字段', 'HVFHV 独有', '乘客发起叫车请求的时间'),
    ('on_scene_datetime', '司机到达时间', 'TIMESTAMP', '时间字段', 'HVFHV 独有', '司机到达接客地点的时间'),
    ('shared_request_flag', '请求共享标志', 'VARCHAR(1)', '标志位', 'HVFHV 独有，Y/N', '是否请求共享出行'),
    ('shared_match_flag', '实际共享标志', 'VARCHAR(1)', '标志位', 'HVFHV 独有，Y/N', '是否实际匹配到共享'),
    ('wav_request_flag', '请求WAV标志', 'VARCHAR(1)', '标志位', 'HVFHV 独有，Y/N', '是否请求无障碍车辆'),
    ('wav_match_flag', '匹配WAV标志', 'VARCHAR(1)', '标志位', 'HVFHV 独有，Y/N', '是否匹配到无障碍车辆'),
    ('access_a_ride_flag', 'MTA无障碍标志', 'VARCHAR(1)', '标志位', 'HVFHV 独有，Y/N', '是否为 MTA 无障碍出行'),
    # 质量标记
    ('is_time_anomaly', '是否时间异常', 'BOOLEAN', '质量标记', 'pickup_at 不在 2026Q1 范围，或 dropoff < pickup，或时长>6h', ''),
    ('is_location_missing', '是否位置缺失', 'BOOLEAN', '质量标记', 'pickup_location_id 或 dropoff_location_id IS NULL', 'FHV 类大面积标记'),
    ('is_distance_outlier', '是否距离异常', 'BOOLEAN', '质量标记', 'distance_miles IS NULL 或 ≤0 或 >500', ''),
    ('source_row_hash', '来源行哈希', 'VARCHAR(64)', '溯源字段', 'MD5(trip_source + 原始行所有字段拼接)', '用于数据血缘追踪'),
    ('source_table', '来源表', 'VARCHAR', '溯源字段', 'Bronze 来源表名', '如 yellow_tripdata_2026q1'),
]
make_sheet(wb, 'trip_detail', 'silver.trip_detail', '行程明细标准表', trip_fields)

# vehicle_detail
vehicle_fields = [
    ('vehicle_id', '车辆代理主键', 'BIGINT', '主键', '自增代理键', 'Silver 层生成，因三表来源不同无法用单一自然键'),
    ('license_number', '牌照编号', 'VARCHAR', '候选键', 'TLC 牌照唯一编号，来源于 License Number / Vehicle License Number', '不允许为空，三表合并键'),
    ('license_type', '牌照类型', 'VARCHAR', '维度属性', 'For Hire Vehicle / Medallion / Stand By Vehicle / Paratransit / Commuter Van', '来自 active_vehicles，分类最全'),
    ('license_status', '牌照状态', 'VARCHAR', '维度属性', 'Current / Suspended', '来自 active_vehicles'),
    ('owner_name', '车主姓名/公司名', 'VARCHAR', '维度属性', '车辆所有权人名称', '来自 active_vehicles'),
    ('expiration_date', '牌照到期日期', 'DATE', '时间字段', '牌照有效截止日期', 'VARCHAR→DATE。早于 2026-01-01 标记已过期'),
    ('dmv_plate_number', 'DMV车牌号', 'VARCHAR', '维度属性', '纽约州 DMV 登记的车牌号码', '来自 active_vehicles，覆盖最全'),
    ('vin', '车辆识别码', 'VARCHAR', '维度属性', '17 位车辆识别代号', '来自 active_vehicles，格式最规范'),
    ('vehicle_make', '车辆品牌', 'VARCHAR', '维度属性', '制造商/品牌名称，如 Toyota、Honda', '来自 active_vehicles'),
    ('vehicle_model', '车辆型号', 'VARCHAR', '维度属性', '具体车型，如 Camry、Suburban', '来自 active_vehicles'),
    ('vehicle_year', '车辆年份', 'INTEGER', '维度属性', '出厂年份 2011-2026', 'VARCHAR→INTEGER。<2000 或 >2027 标记异常'),
    ('fuel_type', '燃料类型', 'VARCHAR', '维度属性', 'Gasoline/Hybrid/Electric/Flex/Plug-in/Diesel/Bio（7种）', '来自 active_vehicles，分类最细'),
    ('wav_flag', '无障碍车辆标志', 'VARCHAR', '维度属性', 'YES/NO/Pilot', '来自 active_vehicles（缺失率仅2.4%）。为空时填 UNKNOWN'),
    ('stretch_limo', '是否加长豪华轿车', 'VARCHAR', '维度属性', 'YES/NO', 'active_vehicles 独有'),
    ('medallion_type', 'Medallion类型', 'VARCHAR', '维度属性', 'Owner must driver / Named Driver', 'medallion_authorized_vehicles 独有'),
    ('base_number', '基地编号', 'VARCHAR', '维度外键', '所属运营基地编号', '来自 fhv_active_vehicles，FHV 专有'),
    ('base_name', '基地名称', 'VARCHAR', '维度属性', '基地官方名称', '来自 fhv_active_vehicles，FHV 专有'),
    ('base_type', '基地类型', 'VARCHAR', '维度属性', 'BLACK CAR / LUXURY / LIVERY 等', '来自 fhv_active_vehicles'),
    ('base_address', '基地地址', 'VARCHAR', '维度属性', '基地物理地址', '来自 fhv_active_vehicles'),
    ('agent_number', '代理编号', 'VARCHAR', '维度外键', 'Medallion 管理代理编号', 'medallion_authorized_vehicles 独有'),
    ('agent_name', '代理名称', 'VARCHAR', '维度属性', '代理机构名称', '缺失率 47%，medallion 独有'),
    ('insurance_carrier', '保险公司名称', 'VARCHAR', '维度属性', '承保公司名称', 'active_vehicles 独有'),
    ('insurance_policy_number', '保险单号', 'VARCHAR', '维度属性', '保单唯一编号', 'active_vehicles 独有'),
    ('last_date_updated', '最后更新日期', 'DATE', '时间字段', '记录最后更新日期', 'VARCHAR→DATE'),
    ('source_table', '来源表', 'VARCHAR', '溯源字段', '记录数据来自哪张 Bronze 表', 'active_vehicles / fhv_active_vehicles / medallion_authorized_vehicles'),
]
make_sheet(wb, 'vehicle_detail', 'silver.vehicle_detail', '车辆明细标准表', vehicle_fields)

# driver_detail
driver_fields = [
    ('driver_id', '司机代理主键', 'BIGINT', '主键', '自增代理键', '因两表来源不同，使用代理键统一'),
    ('license_number', '司机牌照号', 'VARCHAR', '候选键', 'TLC 颁发的司机牌照编号', '不允许为空。同一人可能同时持有 FHV 和 SHL 牌照，需配合 driver_type 唯一'),
    ('driver_name', '司机姓名', 'VARCHAR', '维度属性', '司机全名', ''),
    ('driver_type', '司机类型', 'VARCHAR', '候选键', 'FOR HIRE VEHICLE DRIVER / SHL DRIVER', '与 license_number 组成复合候选键'),
    ('status_code', '状态码', 'INTEGER', '维度属性', 'SHL 独有，1/2/3 许可级别', 'VARCHAR→INTEGER。FHV 司机为 NULL'),
    ('status_description', '状态描述', 'VARCHAR', '维度属性', '状态码的文字说明', 'SHL 独有'),
    ('expiration_date', '牌照到期日期', 'DATE', '时间字段', '牌照有效截止日期', 'VARCHAR→DATE。早于 2026-01-01 标记已过期'),
    ('wav_trained', '是否WAV培训', 'VARCHAR', '维度属性', 'WAV 或空', 'FHV 独有。表示司机是否完成无障碍车辆培训'),
    ('last_date_updated', '最后更新日期', 'DATE', '时间字段', '数据更新日期', 'VARCHAR→DATE'),
    ('last_time_updated', '最后更新时间', 'VARCHAR', '时间字段', 'HH:MM:SS 格式', '保留为 VARCHAR，不精确到秒的查询场景够用'),
    ('source_table', '来源表', 'VARCHAR', '溯源字段', 'fhv_active_drivers / shl_active_drivers', ''),
]
make_sheet(wb, 'driver_detail', 'silver.driver_detail', '司机明细标准表', driver_fields)

# base_detail
base_fields = [
    ('base_detail_id', '基地明细代理主键', 'BIGINT', '主键', '自增代理键', '因 Base License Number 非唯一（1,117 vs 58,923），必须用代理键'),
    ('base_license_number', '基地牌照号', 'VARCHAR', '维度外键', '1,117 个唯一值', '与 year+month 组成复合键才能唯一标识一行'),
    ('base_name', '基地名称', 'VARCHAR', '维度属性', '基地官方名称', ''),
    ('dba', '经营别名', 'VARCHAR', '维度属性', 'Doing Business As，基地的对外经营名称', '缺失率 80.7%。与原表 Base Name 可能不同，保留用于 Agent 模糊匹配基地名称'),
    ('year', '年份', 'INTEGER', '时间字段', '数据所属年份', 'VARCHAR→INTEGER'),
    ('month', '月份', 'INTEGER', '时间字段', '数据所属月份 1-12', 'VARCHAR→INTEGER'),
    ('month_name', '月份名称', 'VARCHAR', '维度属性', 'January-December 英文名', '保留便于可读性和按月份名称排序'),
    ('total_dispatched_trips', '调度行程总数', 'BIGINT', '度量', '该基地当月调度的总行程数', 'VARCHAR→BIGINT。原表全部 VARCHAR 需强制转换'),
    ('total_dispatched_shared_trips', '共享行程数', 'BIGINT', '度量', '其中共享出行的行程数', 'VARCHAR→BIGINT'),
    ('unique_dispatched_vehicles', '去重调度车辆数', 'BIGINT', '度量', '该基地当月调度的去重车辆数', 'VARCHAR→BIGINT'),
    ('composite_key', '复合键', 'VARCHAR', '候选键', 'base_license_number + _ + year + _ + month', '用于去重和关联的唯一标识'),
    ('is_duplicate_key', '是否复合键重复', 'BOOLEAN', '质量标记', '复合键出现 > 1 次时为 TRUE', '标记同一基地同月的多条记录'),
]
make_sheet(wb, 'base_detail', 'silver.base_detail', '基地月度明细标准表', base_fields)

# driver_application_detail
app_fields = [
    ('application_id', '申请代理主键', 'BIGINT', '主键', '自增代理键', '统一主键格式'),
    ('app_no', '申请编号', 'VARCHAR', '候选键', '原始主键，格式如 HDR001234', '不允许为空，必须唯一'),
    ('application_type', '申请类型', 'VARCHAR', '维度属性', 'HDR：Medallion/FHV 司机', ''),
    ('app_date', '申请日期', 'DATE', '时间字段', '提交申请日期', 'VARCHAR→DATE'),
    ('status', '审批状态', 'VARCHAR', '维度属性', 'Incomplete / Pending Fitness Interview / Approved / Denied', '用于司机申请通过率分析'),
    ('fru_interview_scheduled', '体能审查面试状态', 'VARCHAR', '维度属性', '面试安排状态', 'FRU = Fitness Review Unit'),
    ('drug_test', '药检状态', 'VARCHAR', '维度属性', 'Needed / Passed / Not Applicable', ''),
    ('wav_course', 'WAV培训状态', 'VARCHAR', '维度属性', '无障碍车辆培训完成状态', ''),
    ('defensive_driving', '防御性驾驶状态', 'VARCHAR', '维度属性', '6 小时 NYS 防御性驾驶课程完成状态', ''),
    ('driver_exam', '司机考试状态', 'VARCHAR', '维度属性', 'TLC 司机考试通过状态', ''),
    ('medical_clearance_form', '体检表状态', 'VARCHAR', '维度属性', '医疗许可表提交状态', ''),
    ('other_requirements', '其他要求状态', 'VARCHAR', '维度属性', '其他审批材料状态', ''),
    ('last_updated', '最后更新日期', 'DATE', '时间字段', '记录最后更新日期', 'VARCHAR→DATE'),
    ('source_table', '来源表', 'VARCHAR', '溯源字段', 'new_driver_applications', ''),
]
make_sheet(wb, 'driver_application_detail', 'silver.driver_application_detail', '司机申请明细标准表', app_fields)

# parking_violation_detail
parking_fields = [
    ('violation_id', '罚单代理主键', 'BIGINT', '主键', '自增代理键', '统一主键格式'),
    ('summons_number', '罚单编号', 'VARCHAR', '候选键', '罚单唯一编号', '需在 Silver 层验证唯一性'),
    ('plate_id', '车牌号', 'VARCHAR', '维度属性', '违章车辆车牌号', ''),
    ('registration_state', '注册州', 'VARCHAR', '维度属性', '车辆注册所在州，如 NY', ''),
    ('plate_type', '车牌类型', 'VARCHAR', '维度属性', '45 种车牌类型编码', ''),
    ('issue_date', '开票日期', 'DATE', '时间字段', '罚单开具日期', 'VARCHAR→DATE。不在 2025-07-01~2026-06-30 范围内标记'),
    ('violation_code', '违章代码', 'VARCHAR', '维度外键', '219 种违章代码。金额需在 Gold 层通过此代码关联 dim_violation_type 获取', '不允许为空，否则无法关联金额'),
    ('vehicle_body_type', '车身类型', 'VARCHAR', '维度属性', 'Sedan / SUV 等', ''),
    ('vehicle_make', '车辆品牌', 'VARCHAR', '维度属性', '罚单上手写品牌', ''),
    ('issuing_agency', '开票机构', 'VARCHAR', '维度属性', '执法机构代码', ''),
    ('street_code1', '街道代码1', 'VARCHAR', '维度属性', '地理街道编码，第1级', ''),
    ('street_code2', '街道代码2', 'VARCHAR', '维度属性', '地理街道编码，第2级', ''),
    ('street_code3', '街道代码3', 'VARCHAR', '维度属性', '地理街道编码，第3级', ''),
    ('vehicle_expiration_date', '车辆注册到期日', 'DATE', '时间字段', '车辆注册有效期截止日', 'VARCHAR→DATE'),
    ('violation_precinct', '违章管辖区', 'VARCHAR', '维度属性', '违章发生所在警区编号', ''),
    ('issuer_precinct', '开票管辖区', 'VARCHAR', '维度属性', '开票警员所在警区编号', ''),
    ('issuer_code', '开票人员代码', 'VARCHAR', '维度属性', '执法人员编号', ''),
    ('violation_time', '违章时间', 'VARCHAR', '时间字段', '违章发生的具体时刻', '保留 VARCHAR，时间格式不统一'),
    ('violation_county', '违章所在县', 'VARCHAR', '维度属性', '如 NY（纽约县）、K（国王县/布鲁克林）', ''),
    ('street_name', '街道名称', 'VARCHAR', '维度属性', '违章地点所在街道', ''),
    ('intersecting_street', '交叉街道', 'VARCHAR', '维度属性', '与违章街道交叉的街道名', '缺失率 48.7%'),
    ('date_first_observed', '首次观察日期', 'VARCHAR', '时间字段', '首次发现违章的日期', '保留 VARCHAR'),
    ('law_section', '法律条款', 'VARCHAR', '维度属性', '违章对应的法律条款编号', ''),
    ('sub_division', '法律子条款', 'VARCHAR', '维度属性', '法律子条款编号', ''),
    ('violation_legal_code', '违章法律代码', 'VARCHAR', '维度属性', '法律代码', '缺失率 57.2%'),
    ('vehicle_color', '车辆颜色', 'VARCHAR', '维度属性', '罚单记录颜色', ''),
    ('vehicle_year', '车辆年份', 'INTEGER', '维度属性', '违章车辆出厂年份', 'VARCHAR→INTEGER'),
    ('feet_from_curb', '距路缘英尺数', 'VARCHAR', '度量', '车辆距路缘距离', '保留 VARCHAR，含非数字值'),
    ('violation_description', '违章描述', 'VARCHAR', '维度属性', '违章行为文字描述', ''),
    ('fiscal_year', '财年', 'INTEGER', '时间字段', 'NYC 财年（7月1日-6月30日）', 'VARCHAR→INTEGER'),
    # 质量标记
    ('is_duplicate_summons', '是否重复罚单', 'BOOLEAN', '质量标记', 'summons_number 出现 > 1 次时为 TRUE', ''),
    ('source_row_hash', '来源行哈希', 'VARCHAR(64)', '溯源字段', 'MD5 溯源', '用于数据血缘追踪'),
]
make_sheet(wb, 'parking_violation_detail', 'silver.parking_violation_detail', '停车罚单明细标准表', parking_fields)

# tif_payment_detail
tif_fields = [
    ('payment_id', '支付代理主键', 'BIGINT', '主键', '自增代理键', '因 License Number 非唯一（6,115 vs 48,431），必须用代理键'),
    ('license_number', '牌照号', 'VARCHAR', '维度外键', 'Medallion 牌照编号', '与 payment_date 组成复合键'),
    ('agent_number', '代理编号', 'VARCHAR', '维度属性', '管理代理编号', '缺失率 47.6%'),
    ('hackup_payment_amount', '改装支付金额', 'DECIMAL(12,2)', '金额', 'WAV 车辆改装成功后的支付金额', 'VARCHAR→DECIMAL'),
    ('operational_payment_amount', '运营支付金额', 'DECIMAL(12,2)', '金额', '三年度检查成功后的支付（$1,333/次）', 'VARCHAR→DECIMAL'),
    ('total_payment_amount', '总支付金额', 'DECIMAL(12,2)', '金额', '改装 + 运营支付合计', 'VARCHAR→DECIMAL'),
    ('payment_date', '支付日期', 'DATE', '时间字段', 'ACH 转账日期', 'VARCHAR→DATE'),
    ('last_date_updated', '最后更新日期', 'DATE', '时间字段', '数据更新日期', 'VARCHAR→DATE'),
    ('last_time_updated', '最后更新时间', 'VARCHAR', '时间字段', 'HH:MM:SS 格式', '保留 VARCHAR'),
    ('composite_key', '复合键', 'VARCHAR', '候选键', 'license_number + _ + payment_date', '用于去重'),
    ('is_duplicate_key', '是否复合键重复', 'BOOLEAN', '质量标记', '复合键出现 > 1 次时为 TRUE', ''),
]
make_sheet(wb, 'tif_payment_detail', 'silver.tif_payment_detail', 'TIF支付明细标准表', tif_fields)

# crash_detail
crash_fields = [
    ('crash_id', '事故代理主键', 'BIGINT', '主键', '自增代理键', '统一主键格式'),
    ('collision_id', '事故编号', 'BIGINT', '候选键', '原始事故编号', 'VARCHAR→BIGINT。需验证唯一性'),
    ('crash_at', '事故时间', 'TIMESTAMP', '时间字段', 'crash_date + crash_time 合并', '不允许为空。VARCHAR 合并→TIMESTAMP'),
    ('borough', '行政区', 'VARCHAR', '维度属性', '事故所在行政区', '缺失率 30.4%'),
    ('zip_code', '邮政编码', 'VARCHAR', '维度属性', '事故地点邮编', '缺失率 30.4%'),
    ('latitude', '纬度', 'DOUBLE', '空间字段', 'WGS84 坐标系', 'VARCHAR→DOUBLE。不能为 0'),
    ('longitude', '经度', 'DOUBLE', '空间字段', 'WGS84 坐标系', 'VARCHAR→DOUBLE。不能为 0'),
    ('on_street_name', '所在街道', 'VARCHAR', '维度属性', '事故发生所在街道名称', ''),
    ('cross_street_name', '交叉街道', 'VARCHAR', '维度属性', '最近的交叉街道', '缺失率 82.0%'),
    ('off_street_name', '非街道地址', 'VARCHAR', '维度属性', '非街道地点描述（如停车场）', '缺失率 38.4%'),
    # 伤亡统计
    ('persons_injured', '受伤总人数', 'INTEGER', '度量', '事故涉及的受伤人员总数', 'VARCHAR→INTEGER。<0 或 >100 标记异常'),
    ('persons_killed', '死亡总人数', 'INTEGER', '度量', '事故涉及的死亡人员总数', 'VARCHAR→INTEGER'),
    ('pedestrians_injured', '行人受伤数', 'INTEGER', '度量', '受伤行人数量', 'VARCHAR→INTEGER'),
    ('pedestrians_killed', '行人死亡数', 'INTEGER', '度量', '死亡行人数量', 'VARCHAR→INTEGER'),
    ('cyclist_injured', '骑行者受伤数', 'INTEGER', '度量', '受伤骑行者数量', 'VARCHAR→INTEGER'),
    ('cyclist_killed', '骑行者死亡数', 'INTEGER', '度量', '死亡骑行者数量', 'VARCHAR→INTEGER'),
    ('motorist_injured', '驾驶员受伤数', 'INTEGER', '度量', '受伤机动车驾驶员数量', 'VARCHAR→INTEGER'),
    ('motorist_killed', '驾驶员死亡数', 'INTEGER', '度量', '死亡机动车驾驶员数量', 'VARCHAR→INTEGER'),
    # 涉事车辆（仅保留1-2，3-5缺失率93-99%已弃用）
    ('vehicle_type_1', '涉事车辆1类型', 'VARCHAR', '维度属性', '碰撞中第1辆车的类型代码', '车辆3-5缺失率93-99%，已弃用'),
    ('vehicle_type_2', '涉事车辆2类型', 'VARCHAR', '维度属性', '碰撞中第2辆车的类型代码', ''),
    ('contributing_factor_1', '车辆1事故因素', 'VARCHAR', '维度属性', '导致碰撞的因素（车辆1）', ''),
    ('contributing_factor_2', '车辆2事故因素', 'VARCHAR', '维度属性', '导致碰撞的因素（车辆2）', ''),
    # 质量标记
    ('is_duplicate_collision', '是否重复事故', 'BOOLEAN', '质量标记', 'collision_id 出现 > 1 次时为 TRUE', ''),
    ('is_location_missing', '是否位置缺失', 'BOOLEAN', '质量标记', 'latitude 或 longitude IS NULL', ''),
    ('source_row_hash', '来源行哈希', 'VARCHAR(64)', '溯源字段', 'MD5 溯源', ''),
]
make_sheet(wb, 'crash_detail', 'silver.crash_detail', '事故明细标准表', crash_fields)

# crash_person_detail
person_fields = [
    ('crash_person_id', '事故人员代理主键', 'BIGINT', '主键', '自增代理键', '统一主键格式'),
    ('unique_id', '人员记录编号', 'BIGINT', '候选键', '原始人员记录唯一编号', 'VARCHAR→BIGINT。需验证唯一性'),
    ('collision_id', '事故编号', 'BIGINT', '维度外键', '关联 silver.crash_detail.collision_id', '需验证与 crash_detail 的覆盖关系'),
    ('crash_date', '事故日期', 'DATE', '时间字段', '事故发生日期', 'VARCHAR→DATE'),
    ('crash_time', '事故时间', 'VARCHAR', '时间字段', 'HH:MM 格式', '保留 VARCHAR'),
    ('person_id', '人员编号', 'VARCHAR', '维度属性', '事故内人员序号', ''),
    # 人员属性
    ('person_type', '人员类型', 'VARCHAR', '维度属性', '驾驶员/乘客/行人/骑行者', '核心分析维度'),
    ('person_injury', '伤害程度', 'VARCHAR', '维度属性', '受伤/死亡等级别', '核心分析指标'),
    ('person_sex', '性别', 'VARCHAR', '维度属性', '性别标识', ''),
    ('person_age', '年龄', 'INTEGER', '度量', '涉事人员年龄', 'VARCHAR→INTEGER。<0 或 >120 标记异常'),
    ('vehicle_id', '涉事车辆ID', 'VARCHAR', '维度外键', '关联事故中的涉事车辆信息', ''),
    ('ped_role', '行人角色', 'VARCHAR', '维度属性', '行人的具体角色类型', ''),
    # 辅助字段（缺失率46-49%）
    ('ejection', '是否弹出', 'VARCHAR', '维度属性', '人员是否被弹出车外', '缺失率 48.6%'),
    ('emotional_status', '情绪状态', 'VARCHAR', '维度属性', '人员当时情绪状态', '缺失率 46.9%'),
    ('bodily_injury', '身体伤害', 'VARCHAR', '维度属性', '身体伤害描述', '缺失率 46.9%'),
    ('position_in_vehicle', '车内位置', 'VARCHAR', '维度属性', '人员在车内的座位位置', '缺失率 48.6%'),
    ('safety_equipment', '安全设备', 'VARCHAR', '维度属性', '使用的安全设备（安全带/安全座椅等）', '缺失率 48.6%'),
    ('complaint', '投诉信息', 'VARCHAR', '维度属性', '人员投诉内容', '缺失率 46.9%'),
    # 质量标记
    ('is_duplicate_person', '是否重复记录', 'BOOLEAN', '质量标记', 'unique_id 出现 > 1 次时为 TRUE', ''),
    ('is_orphan_record', '是否孤立记录', 'BOOLEAN', '质量标记', 'collision_id 在 silver.crash_detail 中不存在时为 TRUE', '当前覆盖不完全'),
    ('has_missing_aux', '是否缺失辅助字段', 'BOOLEAN', '质量标记', '6 个辅助字段全部为 NULL 时为 TRUE', ''),
    ('source_row_hash', '来源行哈希', 'VARCHAR(64)', '溯源字段', 'MD5 溯源', ''),
]
make_sheet(wb, 'crash_person_detail', 'silver.crash_person_detail', '事故人员明细标准表', person_fields)

# 保存
output_path = r'D:\ProgramData\Datawarehouse\纽约市城市交通\分析报告\Silver层数据字典.xlsx'
wb.save(output_path)
print(f'XLSX saved: {output_path}')
print(f'Sheets: {len(wb.sheetnames)}')
for s in wb.sheetnames:
    print(f'  - {s}')
