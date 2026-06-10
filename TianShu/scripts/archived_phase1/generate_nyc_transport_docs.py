from __future__ import annotations

import json
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
import pyarrow.compute as pc
import pyarrow.parquet as pq
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BASE_DIR = Path("D:/Program Files/Datawarehouse/" + "\u7ebd\u7ea6\u5e02\u57ce\u5e02\u4ea4\u901a")
CHUNK_ROOT = Path("D:/PycharmProjects/PySpark/data")
GENERATED_AT = datetime.now().strftime("%Y-%m-%d %H:%M:%S")


FIELD_ZH = {
    "VendorID": "供应商编号",
    "tpep_pickup_datetime": "黄色出租车接客时间",
    "tpep_dropoff_datetime": "黄色出租车送客时间",
    "lpep_pickup_datetime": "绿色出租车接客时间",
    "lpep_dropoff_datetime": "绿色出租车送客时间",
    "pickup_datetime": "接客时间",
    "dropOff_datetime": "送客时间",
    "dropoff_datetime": "送客时间",
    "request_datetime": "乘客请求时间",
    "on_scene_datetime": "车辆到达现场时间",
    "PULocationID": "上车区域编号",
    "DOLocationID": "下车区域编号",
    "PUlocationID": "上车区域编号",
    "DOlocationID": "下车区域编号",
    "passenger_count": "乘客人数",
    "trip_distance": "行程距离",
    "trip_miles": "行程英里数",
    "trip_time": "行程时长秒数",
    "fare_amount": "车费金额",
    "base_passenger_fare": "基础乘客费用",
    "total_amount": "总金额",
    "tip_amount": "小费金额",
    "tips": "小费金额",
    "tolls_amount": "过路费金额",
    "tolls": "过路费金额",
    "driver_pay": "司机收入",
    "payment_type": "支付方式",
    "congestion_surcharge": "拥堵附加费",
    "cbd_congestion_fee": "中央商务区拥堵费",
    "Airport_fee": "机场费",
    "airport_fee": "机场费",
    "hvfhs_license_num": "高流量网约车牌照号",
    "dispatching_base_num": "派车基地编号",
    "originating_base_num": "原始基地编号",
    "Affiliated_base_number": "关联基地编号",
    "SR_Flag": "共享乘车标记",
    "shared_request_flag": "请求共享标记",
    "shared_match_flag": "成功拼车标记",
    "access_a_ride_flag": "无障碍接驳标记",
    "wav_request_flag": "无障碍车辆请求标记",
    "wav_match_flag": "无障碍车辆匹配标记",
    "LocationID": "出租车区域编号",
    "Borough": "行政区",
    "Zone": "出租车区域",
    "service_zone": "服务区域",
    "Active": "是否活跃",
    "Vehicle License Number": "车辆牌照号",
    "License Number": "牌照号",
    "License Type": "牌照类型",
    "Expiration Date": "到期日期",
    "DMV License Plate Number": "DMV车牌号",
    "DMV Plate Number": "DMV车牌号",
    "Vehicle VIN Number": "车辆VIN码",
    "VIN": "车辆VIN码",
    "Vehicle Year": "车辆年份",
    "Vehicle Make": "车辆品牌",
    "Vehicle Model": "车辆型号",
    "Fuel Type": "燃料类型",
    "Wheelchair Accessible": "是否无障碍车辆",
    "Wheelchair Accessible Trained": "是否完成无障碍培训",
    "WAV": "无障碍车辆标记",
    "Base Number": "基地编号",
    "Base Name": "基地名称",
    "Base Type": "基地类型",
    "Base Address": "基地地址",
    "Owner Name": "车主名称",
    "TLC Vehicle License Status": "TLC车辆牌照状态",
    "Affiliated Base/ Agent Number": "关联基地或代理编号",
    "Current Status": "当前状态",
    "Medallion Type": "出租车牌照类型",
    "Agent Number": "代理编号",
    "Agent Name": "代理名称",
    "App No": "申请编号",
    "Type": "类型",
    "App Date": "申请日期",
    "Status": "状态",
    "Last Updated": "最后更新时间",
    "Last Date Updated": "最后更新日期",
    "Last Time Updated": "最后更新时间",
    "Payment Date": "支付日期",
    "Total Payment Amount": "总支付金额",
    "Hackup Payment Amount": "上牌改装支付金额",
    "Operational Payment Amount": "运营支付金额",
    "crash_date": "事故日期",
    "crash_time": "事故时间",
    "collision_id": "事故编号",
    "borough": "行政区",
    "zip_code": "邮编",
    "latitude": "纬度",
    "longitude": "经度",
    "location": "地理位置",
    "on_street_name": "所在街道",
    "cross_street_name": "交叉街道",
    "off_street_name": "相邻街道",
    "number_of_persons_injured": "受伤人数",
    "number_of_persons_killed": "死亡人数",
    "number_of_pedestrians_injured": "行人受伤人数",
    "number_of_pedestrians_killed": "行人死亡人数",
    "number_of_cyclist_injured": "骑行者受伤人数",
    "number_of_cyclist_killed": "骑行者死亡人数",
    "number_of_motorist_injured": "机动车乘员受伤人数",
    "number_of_motorist_killed": "机动车乘员死亡人数",
    "contributing_factor_vehicle_1": "车辆1碰撞因素",
    "vehicle_type_code1": "车辆1类型",
    "unique_id": "人员记录编号",
    "person_id": "人员编号",
    "person_type": "人员类型",
    "person_injury": "人员伤害程度",
    "vehicle_id": "事故车辆编号",
    "person_age": "人员年龄",
    "person_sex": "人员性别",
    "summons_number": "罚单编号",
    "plate_id": "车牌号",
    "registration_state": "注册州",
    "plate_type": "车牌类型",
    "issue_date": "开票日期",
    "violation_code": "违章代码",
    "violation_description": "违章描述",
    "violation_time": "违章时间",
    "street_name": "街道名称",
    "violation_county": "违章县区",
    "fiscal_year": "财年",
}


def zh_field(name: str) -> str:
    return FIELD_ZH.get(name, f"待补充：{name}")


def type_label(raw_type: str) -> str:
    lower = raw_type.lower()
    if "timestamp" in lower or "date" in lower:
        return "日期时间"
    if any(item in lower for item in ["int", "double", "float", "decimal"]):
        return "数值"
    if "struct" in lower:
        return "结构对象"
    return "文本"


def dataset_domain(name: str, folder: str) -> str:
    text = f"{name} {folder}".lower()
    if any(key in text for key in ["tripdata", "taxi"]):
        return "出行域"
    if any(key in text for key in ["active_vehicles", "vehicles", "medallion"]):
        return "资产域"
    if any(key in text for key in ["drivers", "driver_application", "司机", "base"]):
        return "供给域"
    if any(key in text for key in ["crash", "collision", "事故"]):
        return "安全域"
    if any(key in text for key in ["parking", "violation", "tif", "authorized", "application", "罚单", "授权"]):
        return "监管合规域"
    if any(key in text for key in ["zone", "lookup", "shp"]):
        return "空间地理域"
    return "待归类"


def dataset_role(name: str, folder: str) -> str:
    text = f"{name} {folder}".lower()
    if "tripdata" in text:
        return "行程事实表"
    if "taxi_zone_lookup" in text:
        return "空间维表"
    if "active_vehicles" in text or "vehicles" in text:
        return "车辆/牌照维表"
    if "drivers" in text or "司机" in text:
        return "司机快照表"
    if "application" in text:
        return "申请状态事实表"
    if "tif" in text:
        return "支付事实表"
    if "parking" in text:
        return "罚单事实表"
    if "crash_person" in text:
        return "事故人员明细表"
    if "crash" in text or "collision" in text:
        return "事故事实表"
    return "参考表"


def relation_key(columns: list[str]) -> str:
    candidates = []
    for col in columns:
        low = col.lower()
        if any(key in low for key in ["locationid", "collision_id", "license", "plate", "vin", "base", "summons", "app no"]):
            candidates.append(col)
    return " / ".join(candidates[:6]) if candidates else "需在标准层补充或生成"


def primary_key_hint(name: str, columns: list[str]) -> str:
    for key in ["collision_id", "unique_id", "summons_number", "LocationID", "App No", "Vehicle License Number", "License Number"]:
        if key in columns:
            return key
    if "tripdata" in name:
        return "无自然主键，建议生成 trip_id"
    return "需确认"


def read_csv_header(path: Path) -> tuple[list[str], int]:
    header = list(pd.read_csv(path, nrows=0, encoding="utf-8-sig").columns)
    with path.open("r", encoding="utf-8-sig", errors="replace") as file:
        rows = max(sum(1 for _ in file) - 1, 0)
    return header, rows


def missing_stats_csv(path: Path, columns: list[str]) -> dict[str, dict[str, Any]]:
    df = pd.read_csv(path, encoding="utf-8-sig", dtype=str)
    stats = {}
    for col in columns:
        series = df[col]
        stats[col] = {
            "缺失数": int(series.isna().sum()),
            "缺失率": float(series.isna().mean()),
            "唯一值数": int(series.nunique(dropna=True)),
            "非空数": int(series.notna().sum()),
        }
    return stats


def collect_main_metadata() -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    summaries: list[dict[str, Any]] = []
    fields: list[dict[str, Any]] = []
    quality: list[dict[str, Any]] = []

    for path in sorted(BASE_DIR.rglob("*.parquet")):
        parquet_file = pq.ParquetFile(path)
        columns = [field.name for field in parquet_file.schema_arrow]
        summaries.append(
            {
                "数据集": path.stem,
                "目录": path.parent.name,
                "数据域": dataset_domain(path.name, path.parent.name),
                "文件格式": "parquet",
                "行数": parquet_file.metadata.num_rows,
                "列数": len(columns),
                "大小MB": round(path.stat().st_size / 1024 / 1024, 2),
                "数据角色": dataset_role(path.name, path.parent.name),
                "建议主键": primary_key_hint(path.name, columns),
                "主要关联键": relation_key(columns),
                "完整性状态": "结构完整",
                "备注": "同车型跨月 schema 稳定" if "tripdata" in path.name else "",
            }
        )
        for index, field in enumerate(parquet_file.schema_arrow, 1):
            fields.append(
                {
                    "数据集": path.stem,
                    "序号": index,
                    "字段名": field.name,
                    "中文字段名": zh_field(field.name),
                    "标准类型": type_label(str(field.type)),
                    "原始类型": str(field.type),
                    "字段层级": "关联键" if field.name.lower().endswith("locationid") or "base" in field.name.lower() else "分析字段",
                    "业务说明": zh_field(field.name),
                    "治理备注": "",
                }
            )

    for path in sorted(BASE_DIR.rglob("*.csv")):
        columns, rows = read_csv_header(path)
        summaries.append(
            {
                "数据集": path.stem,
                "目录": path.parent.name,
                "数据域": dataset_domain(path.name, path.parent.name),
                "文件格式": "csv",
                "行数": rows,
                "列数": len(columns),
                "大小MB": round(path.stat().st_size / 1024 / 1024, 2),
                "数据角色": dataset_role(path.name, path.parent.name),
                "建议主键": primary_key_hint(path.name, columns),
                "主要关联键": relation_key(columns),
                "完整性状态": "结构完整",
                "备注": "",
            }
        )
        stats = missing_stats_csv(path, columns)
        for index, col in enumerate(columns, 1):
            field_stats = stats[col]
            note = ""
            if field_stats["缺失率"] >= 0.9:
                note = "高缺失字段，需标注适用范围"
            elif field_stats["缺失率"] >= 0.3:
                note = "缺失率偏高，分析时需谨慎"
            fields.append(
                {
                    "数据集": path.stem,
                    "序号": index,
                    "字段名": col,
                    "中文字段名": zh_field(col),
                    "标准类型": "文本/待推断",
                    "原始类型": "csv文本",
                    "字段层级": "主键/关联键" if col == primary_key_hint(path.name, columns) or col in relation_key(columns) else "分析字段",
                    "业务说明": zh_field(col),
                    "治理备注": note,
                }
            )
            if note:
                quality.append(
                    {
                        "数据集": path.stem,
                        "严重度": "中" if field_stats["缺失率"] < 0.9 else "高",
                        "问题类型": "字段缺失",
                        "问题描述": f"{col} 缺失率约 {field_stats['缺失率']:.1%}",
                        "影响字段": col,
                        "建议处理": note,
                    }
                )

    return summaries, fields, quality


def collect_chunk_metadata() -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    summaries: list[dict[str, Any]] = []
    fields: list[dict[str, Any]] = []
    quality: list[dict[str, Any]] = []

    for folder in ["crash_chunks", "crash_person_chunks", "parking_violations_chunks"]:
        root = CHUNK_ROOT / folder
        files = sorted(root.rglob("*.parquet"))
        if not files:
            continue
        union_columns: list[str] = []
        type_variants: dict[str, set[str]] = defaultdict(set)
        total_rows = 0
        schema_orders = set()
        for path in files:
            parquet_file = pq.ParquetFile(path)
            total_rows += parquet_file.metadata.num_rows
            schema_orders.add(tuple(field.name for field in parquet_file.schema_arrow))
            for field in parquet_file.schema_arrow:
                if field.name not in union_columns:
                    union_columns.append(field.name)
                type_variants[field.name].add(str(field.type))

        summaries.append(
            {
                "数据集": folder,
                "目录": str(root),
                "数据域": dataset_domain(folder, folder),
                "文件格式": "parquet分片",
                "行数": total_rows,
                "列数": len(union_columns),
                "大小MB": round(sum(path.stat().st_size for path in files) / 1024 / 1024, 2),
                "数据角色": dataset_role(folder, folder),
                "建议主键": primary_key_hint(folder, union_columns),
                "主要关联键": relation_key(union_columns),
                "完整性状态": "采集中，不能视为最终全量",
                "备注": f"当前 {len(files)} 个分片，字段顺序变体 {len(schema_orders)} 个",
            }
        )
        for index, col in enumerate(union_columns, 1):
            type_text = " / ".join(sorted(type_variants[col]))
            fields.append(
                {
                    "数据集": folder,
                    "序号": index,
                    "字段名": col,
                    "中文字段名": zh_field(col),
                    "标准类型": type_label(type_text),
                    "原始类型": type_text,
                    "字段层级": "主键/关联键" if col == primary_key_hint(folder, union_columns) or col in relation_key(union_columns) else "分析字段",
                    "业务说明": zh_field(col),
                    "治理备注": "字段存在类型漂移，需统一" if len(type_variants[col]) > 1 else "采集完成后复核",
                }
            )
        if len(schema_orders) > 1:
            quality.append(
                {
                    "数据集": folder,
                    "严重度": "高",
                    "问题类型": "schema漂移",
                    "问题描述": f"分片字段顺序或字段集合存在 {len(schema_orders)} 个变体",
                    "影响字段": "全表",
                    "建议处理": "入湖前按 union schema 固定列顺序，缺失列补空",
                }
            )
        drift = {name: sorted(values) for name, values in type_variants.items() if len(values) > 1}
        if drift:
            quality.append(
                {
                    "数据集": folder,
                    "严重度": "高",
                    "问题类型": "字段类型漂移",
                    "问题描述": json.dumps(drift, ensure_ascii=False),
                    "影响字段": " / ".join(drift.keys()),
                    "建议处理": "标准层统一为稳定结构或拆解为经纬度文本字段",
                }
            )

    return summaries, fields, quality


def date_quality_issues() -> list[dict[str, Any]]:
    issues: list[dict[str, Any]] = []
    date_columns = {
        "yellow_tripdata": ["tpep_pickup_datetime", "tpep_dropoff_datetime"],
        "green_tripdata": ["lpep_pickup_datetime", "lpep_dropoff_datetime"],
        "fhv_tripdata": ["pickup_datetime", "dropOff_datetime"],
        "fhvhv_tripdata": ["request_datetime", "pickup_datetime", "dropoff_datetime"],
    }
    for path in sorted(BASE_DIR.rglob("*.parquet")):
        key = next((item for item in date_columns if path.name.startswith(item)), None)
        if not key:
            continue
        table = pq.read_table(path, columns=date_columns[key])
        for col in date_columns[key]:
            min_value = pc.min(table[col]).as_py()
            max_value = pc.max(table[col]).as_py()
            if min_value.year < 2025 or max_value.year > 2026:
                issues.append(
                    {
                        "数据集": path.stem,
                        "严重度": "中",
                        "问题类型": "异常时间",
                        "问题描述": f"{col} 范围为 {min_value} 至 {max_value}",
                        "影响字段": col,
                        "建议处理": "保留原值并在 Silver 层增加异常时间标记",
                    }
                )
    return issues


def append_sheet(workbook: Workbook, title: str, rows: list[dict[str, Any]] | list[list[Any]]) -> None:
    sheet = workbook.create_sheet(title)
    if not rows:
        sheet.append(["无数据"])
        return
    if isinstance(rows[0], dict):
        headers = list(rows[0].keys())
        sheet.append(headers)
        for row in rows:
            sheet.append([row.get(header, "") for header in headers])
    else:
        for row in rows:
            sheet.append(row)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A2"
    for index, column in enumerate(sheet.columns, 1):
        max_len = min(max(len(str(cell.value)) if cell.value is not None else 0 for cell in column) + 2, 48)
        sheet.column_dimensions[get_column_letter(index)].width = max(max_len, 12)


def build_workbook(
    summaries: list[dict[str, Any]],
    fields: list[dict[str, Any]],
    quality: list[dict[str, Any]],
) -> Path:
    workbook = Workbook()
    workbook.remove(workbook.active)

    domain_rows = [
        {"数据域": "出行域", "中文说明": "描述城市交通出行行为，包括上车、下车、费用、距离、时长", "代表数据": "Yellow/Green/FHV/HVFHV 行程", "建模归属": "fact_trips"},
        {"数据域": "资产域", "中文说明": "描述车辆、牌照、VIN、燃料、无障碍能力等交通资产", "代表数据": "active_vehicles、FHV车辆、Medallion车辆", "建模归属": "dim_vehicle"},
        {"数据域": "供给域", "中文说明": "描述司机、基地、申请人等可服务资源", "代表数据": "FHV司机、SHL司机、新司机申请、Base字段", "建模归属": "dim_driver / dim_base / fact_driver_applications"},
        {"数据域": "安全域", "中文说明": "描述交通事故、伤亡、人员、碰撞因素", "代表数据": "crash_chunks、crash_person_chunks", "建模归属": "fact_crashes / fact_crash_persons"},
        {"数据域": "监管合规域", "中文说明": "描述罚单、牌照状态、车辆授权、补贴支付、申请审批", "代表数据": "parking violations、TIF、授权车辆、申请状态", "建模归属": "fact_parking_violations / fact_tif_payments / dim_license_status"},
        {"数据域": "空间地理域", "中文说明": "描述区域、行政区、经纬度、街道、邮编等空间信息", "代表数据": "taxi_zone_lookup、taxi_zones.shp、经纬度、街道", "建模归属": "dim_taxi_zone"},
    ]
    model_rows = [
        ["模型对象", "中文注释"],
        ["Bronze 原始层", "原始文件保留层，不改字段、不改格式，用于追溯来源"],
        ["Silver 标准层", "清洗标准化层，统一字段名、类型、时间、空间、状态码"],
        ["Gold 主题星型模型", "面向分析和问数的主题模型，事实表围绕公共维表组织"],
        ["fact_trips", "出行事实表，记录出租车、网约车等每一次行程"],
        ["fact_parking_violations", "停车罚单事实表，记录每一张违章罚单"],
        ["fact_tif_payments", "出租车改善基金支付事实表，记录补贴或支付行为"],
        ["fact_driver_applications", "司机申请事实表，记录新司机申请及审批状态"],
        ["dim_vehicle", "车辆维表，描述车辆、VIN、车牌、燃料、WAV等资产属性"],
        ["dim_driver", "司机维表，描述司机牌照、司机类型、有效期等供给属性"],
        ["dim_base", "基地维表，描述FHV基地、基地类型、联系方式和地址"],
        ["dim_taxi_zone", "出租车区域维表，描述LocationID、行政区、服务区域"],
        ["dim_date", "日期维表，统一支持日、周、月、季度、财年等时间分析"],
        ["dim_violation_code", "违章代码维表，解释罚单违章类型和罚款金额"],
        ["dim_license_status", "牌照状态维表，统一解释车辆、司机、申请状态"],
        ["事故 ER 子模型", "安全域保留事件与人员明细的一对多关系"],
        ["fact_crashes", "事故事实表，记录每一次碰撞事故的位置、时间、伤亡和原因"],
        ["fact_crash_persons", "事故人员明细表，记录事故涉及人员、伤害程度和人员属性"],
        ["中文语义层", "面向中文用户和Agent的业务解释层，定义表、字段、指标和关联口径"],
    ]
    semantic_rows = [
        {"对象": "semantic_tables", "中文注释": "中文表说明，解释每张表能回答什么问题"},
        {"对象": "semantic_fields", "中文注释": "中文字段说明，解释字段含义、单位、适用范围"},
        {"对象": "semantic_metrics", "中文注释": "中文指标口径，如行程量、事故率、罚单量、WAV覆盖率"},
        {"对象": "semantic_relationships", "中文注释": "中文关系说明，标注强关联、弱关联和不可关联"},
        {"对象": "semantic_question_templates", "中文注释": "中文问数模板，约束Agent如何安全生成查询"},
    ]
    mapping_rows = []
    for source, standard, comment in [
        ("tpep_pickup_datetime/lpep_pickup_datetime/pickup_datetime", "pickup_at", "统一接客时间"),
        ("tpep_dropoff_datetime/lpep_dropoff_datetime/dropoff_datetime/dropOff_datetime", "dropoff_at", "统一送客时间"),
        ("PULocationID/PUlocationID", "pickup_location_id", "统一上车区域编号"),
        ("DOLocationID/DOlocationID", "dropoff_location_id", "统一下车区域编号"),
        ("License Number/Vehicle License Number", "license_no", "统一牌照号"),
        ("DMV License Plate Number/DMV Plate Number/plate_id", "plate_no", "统一车牌号"),
        ("VIN/Vehicle VIN Number", "vehicle_vin", "统一车辆VIN"),
        ("Base Number/dispatching_base_num/Affiliated Base/ Agent Number", "base_no", "统一基地编号"),
        ("collision_id", "collision_id", "事故主关联键"),
        ("summons_number", "summons_no", "罚单主键"),
    ]:
        mapping_rows.append({"原始字段": source, "标准字段": standard, "中文注释": comment})

    append_sheet(workbook, "数据总览", summaries)
    append_sheet(workbook, "六大数据域", domain_rows)
    append_sheet(workbook, "数据字典", fields)
    append_sheet(workbook, "表关系模型", model_rows)
    append_sheet(workbook, "入湖标准字段映射", mapping_rows)
    append_sheet(workbook, "中文语义层", semantic_rows)
    append_sheet(workbook, "数据质量问题", quality)

    output_path = BASE_DIR / "纽约市城市交通_全域数据规范.xlsx"
    workbook.save(output_path)
    return output_path


def model_tree_text() -> str:
    return """纽约市城市交通数据底座
├─ Bronze 原始层  # 原始文件保留层，不改字段、不改格式，用于追溯来源
├─ Silver 标准层  # 清洗标准化层，统一字段名、类型、时间、空间、状态码
├─ Gold 主题星型模型  # 面向分析和问数的主题模型，事实表围绕公共维表组织
│  ├─ fact_trips  # 出行事实表，记录出租车、网约车等每一次行程
│  ├─ fact_parking_violations  # 停车罚单事实表，记录每一张违章罚单
│  ├─ fact_tif_payments  # 出租车改善基金支付事实表，记录补贴或支付行为
│  ├─ fact_driver_applications  # 司机申请事实表，记录新司机申请及审批状态
│  ├─ dim_vehicle  # 车辆维表，描述车辆、VIN、车牌、燃料、WAV等资产属性
│  ├─ dim_driver  # 司机维表，描述司机牌照、司机类型、有效期等供给属性
│  ├─ dim_base  # 基地维表，描述FHV基地、基地类型、联系方式和地址
│  ├─ dim_taxi_zone  # 出租车区域维表，描述LocationID、行政区、服务区域
│  ├─ dim_date  # 日期维表，统一支持日、周、月、季度、财年等时间分析
│  ├─ dim_violation_code  # 违章代码维表，解释罚单违章类型和罚款金额
│  └─ dim_license_status  # 牌照状态维表，统一解释车辆、司机、申请状态
├─ 事故 ER 子模型  # 安全域保留事件与人员明细的一对多关系
│  ├─ fact_crashes  # 事故事实表，记录每一次碰撞事故的位置、时间、伤亡和原因
│  └─ fact_crash_persons  # 事故人员明细表，记录事故涉及人员、伤害程度和人员属性
└─ 中文语义层  # 面向中文用户和Agent的业务解释层，定义表、字段、指标和关联口径
   ├─ semantic_tables  # 中文表说明，解释每张表能回答什么问题
   ├─ semantic_fields  # 中文字段说明，解释字段含义、单位、适用范围
   ├─ semantic_metrics  # 中文指标口径，如行程量、事故率、罚单量、WAV覆盖率
   ├─ semantic_relationships  # 中文关系说明，标注强关联、弱关联和不可关联
   └─ semantic_question_templates  # 中文问数模板，约束Agent如何安全生成查询"""


def write_markdown_files(summaries: list[dict[str, Any]], quality: list[dict[str, Any]]) -> tuple[Path, Path]:
    domain_counts = defaultdict(int)
    for item in summaries:
        domain_counts[item["数据域"]] += 1
    trip_rows = sum(item["行数"] for item in summaries if item["数据角色"] == "行程事实表")

    feasibility = f"""# 纽约市城市交通 Agent 基座可行性评估

生成时间：{GENERATED_AT}

## 总体结论

当前数据源已经覆盖出行域、资产域、供给域、安全域、监管合规域、空间地理域六大主题，适合建设为城市交通数据分析与数据开发 Agent 的基础底座。

当前阶段评级为 **A-**。下载中分片完成并经过 Silver 标准化后，可提升至 **A**。

## 六大域覆盖

| 数据域 | 当前源表数量 | 评估 |
|---|---:|---|
| 出行域 | {domain_counts.get('出行域', 0)} | TLC 行程数据体量大，结构稳定，是核心事实层 |
| 资产域 | {domain_counts.get('资产域', 0)} | 车辆、牌照、VIN、燃料、WAV 等资产属性较完整 |
| 供给域 | {domain_counts.get('供给域', 0)} | 司机、基地、申请状态具备，但与行程多为弱关联 |
| 安全域 | {domain_counts.get('安全域', 0)} | 事故与人员分片可形成 ER 子模型，当前仍在采集中 |
| 监管合规域 | {domain_counts.get('监管合规域', 0)} | 罚单、授权、TIF、申请审批可形成合规分析主题 |
| 空间地理域 | {domain_counts.get('空间地理域', 0)} | taxi zone、shp、经纬度、街道字段可支撑空间分析 |

## 源表模型判断

源表不是严格 ER 模型，也不是标准星型模型。它们更接近“开放数据宽表集合 + 官方数据字典 + 局部关系表”。

局部接近 ER 的部分是事故数据，`collision_id` 可以连接事故表与事故人员表。局部接近星型的部分是 TLC 行程数据，可以通过 `PULocationID`、`DOLocationID` 关联出租车区域维表。

## 推荐目标模型

```text
{model_tree_text()}
```

## Agent 适配能力

- 支持中文问数：行程量、区域流动、事故分布、罚单趋势、车辆结构、司机供给。
- 支持数据开发：采集分片合并、schema 对齐、分区设计、质量校验、维表构建。
- 支持治理解释：字段中文语义、强弱关联说明、指标口径说明、不可关联限制。

## 主要风险

- 下载中的事故、事故人员、停车罚单分片不能视为最终全量。
- chunk 分片存在字段顺序漂移，事故表 `location` 字段存在结构类型差异。
- TLC 行程表存在少量异常时间，需要在 Silver 层打标或过滤。
- 车辆、司机和行程之间缺少 trip 级强外键，多数只能按基地、牌照、时间快照做弱关联。

## 建设优先级

1. 先建设 `dim_taxi_zone` 和 `fact_trips`。
2. 再建设 `dim_vehicle`、`dim_driver`、`dim_base`。
3. 下载完成后建设 `fact_crashes`、`fact_crash_persons`。
4. 下载完成后建设 `fact_parking_violations` 和 `dim_violation_code`。
5. 最后建设中文语义层，约束 Agent 的中文问数和自动建模行为。
"""

    issue_lines = "\n".join(
        f"| {item['数据集']} | {item['严重度']} | {item['问题类型']} | {item['问题描述']} | {item['建议处理']} |"
        for item in quality[:80]
    )
    completeness = f"""# 纽约市城市交通数据完整性分析报告

生成时间：{GENERATED_AT}

## 总体概况

主目录：`{BASE_DIR}`

当前已识别结构化数据集 {len(summaries)} 个。其中 TLC 行程事实表合计约 {trip_rows:,} 行，是当前最完整、最适合优先建模的数据主题。

## 完整性结论

| 数据域 | 完整性 | 说明 |
|---|---|---|
| 出行域 | 高 | 四类 TLC 行程表跨月 schema 一致，体量充足 |
| 资产域 | 较高 | 车辆、牌照、VIN、燃料、WAV 等字段可用 |
| 供给域 | 中高 | 司机、车辆、基地、申请状态都有，但与行程缺少强外键 |
| 安全域 | 中 | 事故与人员表结构清晰，但分片仍在下载 |
| 监管合规域 | 中 | 罚单分片仍在下载；TIF、授权、申请状态可先纳入规范 |
| 空间地理域 | 高 | taxi zone、shp、经纬度、街道、邮编可作为公共空间支撑 |

## 文件完整性

- 主目录已有 parquet、csv、xlsx、txt、pdf、jpg、shp 等多类文件。
- `.nas_pro_downloading` 属于下载临时文件，不应进入 Bronze 以外的可查询层。
- `crash_chunks`、`crash_person_chunks`、`parking_violations_chunks` 当前仍在采集中，完整性需在下载结束后复核。

## 表结构完整性

- TLC 行程 parquet 各车型跨 2026 年 1-3 月 schema 一致。
- 事故和罚单分片存在字段顺序或字段集合漂移，入湖前必须按 union schema 固定列顺序。
- 事故表 `location` 字段存在结构类型差异，建议在 Silver 层拆为 `latitude`、`longitude`、`location_raw`。

## 关联完整性

- `PULocationID`、`DOLocationID` 可关联 `taxi_zone_lookup.LocationID`。
- `collision_id` 可连接事故表和事故人员表，适合保留 ER 子模型。
- `License Number`、`Vehicle License Number`、`VIN`、`DMV Plate` 适合构建车辆/牌照维度。
- 车辆/司机和行程之间不是 trip 级强关联，Agent 必须避免直接推断“某司机完成某行程”。

## 质量问题清单

| 数据集 | 严重度 | 问题类型 | 问题描述 | 建议处理 |
|---|---|---|---|---|
{issue_lines}

## Agent 使用限制

- 可以回答出行趋势、区域热度、事故分布、罚单趋势、车辆结构、供给规模等问题。
- 不应在缺少强外键时回答单车、单司机、单行程的确定性因果问题。
- 对下载中分片，Agent 只能描述“当前已采集数据”，不能宣称全量覆盖。
- 对异常时间和高缺失字段，Agent 必须在回答中说明过滤口径或数据限制。
"""

    feasibility_path = BASE_DIR / "纽约市城市交通_Agent基座可行性评估.md"
    completeness_path = BASE_DIR / "纽约市城市交通_数据完整性分析报告.md"
    feasibility_path.write_text(feasibility, encoding="utf-8")
    completeness_path.write_text(completeness, encoding="utf-8")
    return feasibility_path, completeness_path


def main() -> None:
    main_summaries, main_fields, main_quality = collect_main_metadata()
    chunk_summaries, chunk_fields, chunk_quality = collect_chunk_metadata()
    quality = main_quality + chunk_quality + date_quality_issues()
    summaries = main_summaries + chunk_summaries
    fields = main_fields + chunk_fields

    workbook_path = build_workbook(summaries, fields, quality)
    feasibility_path, completeness_path = write_markdown_files(summaries, quality)

    print(json.dumps({
        "workbook": str(workbook_path),
        "feasibility": str(feasibility_path),
        "completeness": str(completeness_path),
        "datasets": len(summaries),
        "fields": len(fields),
        "quality_issues": len(quality),
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
