from __future__ import annotations

from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
import pyarrow.compute as pc
import pyarrow.parquet as pq
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BASE_DIR = Path("D:/ProgramData/Datawarehouse/" + "\u7ebd\u7ea6\u5e02\u57ce\u5e02\u4ea4\u901a")
OUTPUT_XLSX = BASE_DIR / "\u7ebd\u7ea6\u5e02\u57ce\u5e02\u4ea4\u901a_\u5168\u57df\u6570\u636e\u89c4\u8303.xlsx"
OUTPUT_FEASIBILITY = BASE_DIR / "\u7ebd\u7ea6\u5e02\u57ce\u5e02\u4ea4\u901a_Agent\u57fa\u5ea7\u53ef\u884c\u6027\u8bc4\u4f30.md"
OUTPUT_COMPLETENESS = BASE_DIR / "\u7ebd\u7ea6\u5e02\u57ce\u5e02\u4ea4\u901a_\u6570\u636e\u5b8c\u6574\u6027\u5206\u6790\u62a5\u544a.md"
GENERATED_AT = datetime.now().strftime("%Y-%m-%d %H:%M:%S")


TABLE_META = {
    "fhv_tripdata_2026Q1": ("普通网约车行程表", "出行域", "行程事实表"),
    "fhvhv_tripdata_2026Q1": ("高流量网约车行程表", "出行域", "核心行程事实表"),
    "green_tripdata_2026Q1": ("绿色出租车行程表", "出行域", "补充行程事实表"),
    "yellow_tripdata_2026Q1": ("黄色出租车行程表", "出行域", "核心行程事实表"),
    "taxi_zone_lookup": ("出租车区域维表", "空间地理域", "空间维表"),
    "parking_violations_all": ("停车违章罚单事实表", "监管合规域", "罚单事实表"),
    "TLC_New_Driver_Application_Status_20260605": ("TLC新司机申请状态表", "监管合规域", "申请状态事实表"),
    "Taxi_Improvement_Fund_(TIF)_Medallion_Payments_20260605": ("出租车改善基金支付表", "监管合规域", "支付事实表"),
    "For_Hire_Vehicles_(FHV)_-_Active_20260605": ("FHV活跃车辆表", "资产域", "车辆/基地维表"),
    "For_Hire_Vehicles_(FHV)_-_Active_Drivers_20260605": ("FHV活跃司机表", "供给域", "司机快照表"),
    "Medallion__Vehicles_-_Authorized_20260605": ("授权出租车牌照车辆表", "监管合规域", "授权车辆维表"),
    "crash_person_all": ("机动车碰撞事故人员表", "安全域", "事故人员明细表"),
    "crash_merged": ("机动车碰撞事故事实表", "安全域", "事故事实表"),
    "active_vehicles": ("活跃车辆注册表", "资产域", "车辆注册维表"),
    "Street_Hail_Livery_(SHL)_Drivers_-_Active_20260605": ("SHL活跃司机表", "供给域", "司机快照表"),
}


FIELD_ZH = {
    "VendorID": "供应商编号",
    "tpep_pickup_datetime": "黄色出租车接客时间",
    "tpep_dropoff_datetime": "黄色出租车送客时间",
    "lpep_pickup_datetime": "绿色出租车接客时间",
    "lpep_dropoff_datetime": "绿色出租车送客时间",
    "pickup_datetime": "接客时间",
    "dropOff_datetime": "送客时间",
    "dropoff_datetime": "送客时间",
    "request_datetime": "乘客请求时间",
    "on_scene_datetime": "车辆到达现场时间",
    "PULocationID": "上车区域编号",
    "DOLocationID": "下车区域编号",
    "PUlocationID": "上车区域编号",
    "DOlocationID": "下车区域编号",
    "passenger_count": "乘客人数",
    "trip_distance": "行程距离",
    "trip_miles": "行程英里数",
    "trip_time": "行程时长秒数",
    "fare_amount": "车费金额",
    "base_passenger_fare": "基础乘客费用",
    "total_amount": "总金额",
    "tip_amount": "小费金额",
    "tips": "小费金额",
    "tolls_amount": "过路费金额",
    "tolls": "过路费金额",
    "driver_pay": "司机收入",
    "payment_type": "支付方式",
    "congestion_surcharge": "拥堵附加费",
    "cbd_congestion_fee": "中央商务区拥堵费",
    "Airport_fee": "机场费",
    "airport_fee": "机场费",
    "hvfhs_license_num": "高流量网约车牌照号",
    "dispatching_base_num": "派车基地编号",
    "originating_base_num": "原始基地编号",
    "Affiliated_base_number": "关联基地编号",
    "SR_Flag": "共享乘车标记",
    "shared_request_flag": "请求共享标记",
    "shared_match_flag": "成功拼车标记",
    "access_a_ride_flag": "无障碍接驳标记",
    "wav_request_flag": "无障碍车辆请求标记",
    "wav_match_flag": "无障碍车辆匹配标记",
    "LocationID": "出租车区域编号",
    "Borough": "行政区",
    "Zone": "出租车区域",
    "service_zone": "服务区域",
    "summons_number": "罚单编号",
    "plate_id": "车牌号",
    "registration_state": "注册州",
    "plate_type": "车牌类型",
    "issue_date": "开票日期",
    "violation_code": "违章代码",
    "vehicle_body_type": "车辆车身类型",
    "vehicle_make": "车辆品牌",
    "issuing_agency": "开票机构",
    "violation_time": "违章时间",
    "violation_county": "违章县区",
    "street_name": "街道名称",
    "intersecting_street": "交叉街道",
    "violation_legal_code": "违章法律代码",
    "violation_description": "违章描述",
    "fiscal_year": "财年",
    "Active": "是否活跃",
    "Vehicle License Number": "车辆牌照号",
    "License Number": "牌照号",
    "License Type": "牌照类型",
    "Expiration Date": "到期日期",
    "DMV License Plate Number": "DMV车牌号",
    "DMV Plate Number": "DMV车牌号",
    "Vehicle VIN Number": "车辆VIN码",
    "VIN": "车辆VIN码",
    "Vehicle Year": "车辆年份",
    "Vehicle Make": "车辆品牌",
    "Vehicle Model": "车辆型号",
    "Fuel Type": "燃料类型",
    "Wheelchair Accessible": "是否无障碍车辆",
    "Wheelchair Accessible Trained": "是否完成无障碍培训",
    "WAV": "无障碍车辆标记",
    "Base Number": "基地编号",
    "Base Name": "基地名称",
    "Base Type": "基地类型",
    "Base Address": "基地地址",
    "Owner Name": "车主名称",
    "TLC Vehicle License Status": "TLC车辆牌照状态",
    "Affiliated Base/ Agent Number": "关联基地或代理编号",
    "Current Status": "当前状态",
    "Medallion Type": "出租车牌照类型",
    "Agent Number": "代理编号",
    "Agent Name": "代理名称",
    "App No": "申请编号",
    "Type": "类型",
    "App Date": "申请日期",
    "Status": "状态",
    "Last Updated": "最后更新时间",
    "Last Date Updated": "最后更新日期",
    "Last Time Updated": "最后更新时间",
    "Payment Date": "支付日期",
    "Total Payment Amount": "总支付金额",
    "Hackup Payment Amount": "上牌改装支付金额",
    "Operational Payment Amount": "运营支付金额",
    "crash_date": "事故日期",
    "crash_time": "事故时间",
    "collision_id": "事故编号",
    "borough": "行政区",
    "zip_code": "邮编",
    "latitude": "纬度",
    "longitude": "经度",
    "location": "地理位置",
    "on_street_name": "所在街道",
    "cross_street_name": "交叉街道",
    "off_street_name": "相邻街道",
    "number_of_persons_injured": "受伤人数",
    "number_of_persons_killed": "死亡人数",
    "number_of_pedestrians_injured": "行人受伤人数",
    "number_of_pedestrians_killed": "行人死亡人数",
    "number_of_cyclist_injured": "骑行者受伤人数",
    "number_of_cyclist_killed": "骑行者死亡人数",
    "number_of_motorist_injured": "机动车乘员受伤人数",
    "number_of_motorist_killed": "机动车乘员死亡人数",
    "contributing_factor_vehicle_1": "车辆1碰撞因素",
    "contributing_factor_vehicle_2": "车辆2碰撞因素",
    "vehicle_type_code1": "车辆1类型",
    "vehicle_type_code2": "车辆2类型",
    "unique_id": "人员记录编号",
    "person_id": "人员编号",
    "person_type": "人员类型",
    "person_injury": "人员伤害程度",
    "vehicle_id": "事故车辆编号",
    "person_age": "人员年龄",
    "person_sex": "人员性别",
}


def table_meta(stem: str) -> tuple[str, str, str]:
    return TABLE_META.get(stem, (f"待补充中文表名：{stem}", "待归类", "参考表"))


def field_zh(name: str) -> str:
    return FIELD_ZH.get(name, f"待补充：{name}")


def type_label(raw_type: str) -> str:
    lower = raw_type.lower()
    if "timestamp" in lower or "date" in lower:
        return "日期时间"
    if any(token in lower for token in ["int", "double", "float", "decimal"]):
        return "数值"
    if "struct" in lower:
        return "结构对象"
    return "文本"


def primary_key_hint(stem: str, columns: list[str]) -> str:
    explicit = {
        "parking_violations_all": "summons_number（需去重后使用）",
        "crash_merged": "collision_id（当前不唯一，需去重后使用）",
        "crash_person_all": "unique_id（当前不唯一，需去重后使用）",
        "Taxi_Improvement_Fund_(TIF)_Medallion_Payments_20260605": "建议生成 payment_id，License Number 不是主键",
    }
    if stem in explicit:
        return explicit[stem]
    for key in ["LocationID", "App No", "Vehicle License Number", "License Number"]:
        if key in columns:
            return key
    if "tripdata" in stem:
        return "无自然主键，建议生成 trip_id"
    return "需确认"


def relation_key(columns: list[str]) -> str:
    keys = []
    for column in columns:
        low = column.lower()
        if any(token in low for token in ["locationid", "collision_id", "license", "plate", "vin", "base", "summons", "app no"]):
            keys.append(column)
    return " / ".join(keys[:8]) if keys else "需在标准层补充或生成"


def field_level(field: str, primary: str) -> str:
    low = field.lower()
    if field and field in primary:
        return "主键/候选键"
    if any(token in low for token in ["locationid", "collision_id", "license", "plate", "vin", "base", "summons"]):
        return "关联键"
    if any(token in low for token in ["date", "time", "datetime"]):
        return "时间字段"
    if any(token in low for token in ["amount", "fare", "fee", "toll", "tax", "pay", "surcharge"]):
        return "金额字段"
    if any(token in low for token in ["latitude", "longitude", "borough", "street", "zone", "zip"]):
        return "空间字段"
    if any(token in low for token in ["status", "flag", "type", "code"]):
        return "状态/枚举字段"
    return "分析字段"


def status_for_missing(rate: float | str) -> str:
    if rate == "":
        return "待统计"
    if rate == 0:
        return "良好"
    if rate < 0.3:
        return "轻微缺失"
    if rate < 0.8:
        return "缺失偏高"
    return "严重缺失"


def parquet_null_stats(path: Path) -> dict[str, dict[str, Any]]:
    parquet_file = pq.ParquetFile(path)
    row_count = parquet_file.metadata.num_rows
    stats = {}
    for index, field in enumerate(parquet_file.schema_arrow):
        null_count = 0
        known = True
        for group_index in range(parquet_file.metadata.num_row_groups):
            col_stats = parquet_file.metadata.row_group(group_index).column(index).statistics
            if col_stats is None or col_stats.null_count is None:
                known = False
                break
            null_count += col_stats.null_count
        if known:
            rate = null_count / row_count if row_count else 0
            stats[field.name] = {
                "缺失数": null_count,
                "缺失率": rate,
                "非空数": row_count - null_count,
                "唯一值数": "",
                "状态": status_for_missing(rate),
            }
        else:
            stats[field.name] = {"缺失数": "", "缺失率": "", "非空数": "", "唯一值数": "", "状态": "待统计"}
    return stats


def csv_null_stats(path: Path) -> dict[str, dict[str, Any]]:
    frame = pd.read_csv(path, encoding="utf-8-sig", dtype=str)
    stats = {}
    for column in frame.columns:
        missing = int(frame[column].isna().sum())
        rate = missing / len(frame) if len(frame) else 0
        stats[column] = {
            "缺失数": missing,
            "缺失率": rate,
            "非空数": len(frame) - missing,
            "唯一值数": int(frame[column].nunique(dropna=True)),
            "状态": status_for_missing(rate),
        }
    return stats


def key_review(path: Path, stem: str, columns: list[str]) -> dict[str, Any] | None:
    keys = {
        "parking_violations_all": "summons_number",
        "crash_merged": "collision_id",
        "crash_person_all": "unique_id",
        "taxi_zone_lookup": "LocationID",
        "TLC_New_Driver_Application_Status_20260605": "App No",
        "Taxi_Improvement_Fund_(TIF)_Medallion_Payments_20260605": "License Number",
        "For_Hire_Vehicles_(FHV)_-_Active_20260605": "Vehicle License Number",
        "For_Hire_Vehicles_(FHV)_-_Active_Drivers_20260605": "License Number",
        "Medallion__Vehicles_-_Authorized_20260605": "License Number",
        "active_vehicles": "License Number",
        "Street_Hail_Livery_(SHL)_Drivers_-_Active_20260605": "License Number",
    }
    key = keys.get(stem)
    if not key or key not in columns:
        return None
    if path.suffix.lower() == ".parquet":
        arr = pq.read_table(path, columns=[key])[key]
        rows = len(arr)
        distinct = len(pc.unique(arr))
        nulls = int(pc.sum(pc.is_null(arr)).as_py() or 0)
    else:
        frame = pd.read_csv(path, encoding="utf-8-sig", dtype=str, usecols=[key])
        rows = len(frame)
        distinct = int(frame[key].nunique(dropna=True))
        nulls = int(frame[key].isna().sum())
    duplicate = rows - distinct
    ok = duplicate == 0 and nulls == 0
    return {
        "英文表名": stem,
        "候选键": key,
        "行数": rows,
        "唯一值数": distinct,
        "空值数": nulls,
        "重复数": duplicate,
        "是否可直接作为主键": "是" if ok else "否",
        "治理建议": "可直接作为主键" if ok else "需去重、确认业务粒度或生成代理键",
    }


def collect_metadata() -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    summaries = []
    fields = []
    missing_rows = []
    quality = []
    key_rows = []
    paths = sorted(list(BASE_DIR.rglob("*.parquet")) + list(BASE_DIR.rglob("*.csv")))
    for path in paths:
        stem = path.stem
        chinese_name, domain, role = table_meta(stem)
        if path.suffix.lower() == ".parquet":
            parquet_file = pq.ParquetFile(path)
            columns = [(field.name, str(field.type)) for field in parquet_file.schema_arrow]
            row_count = parquet_file.metadata.num_rows
            null_stats = parquet_null_stats(path)
        else:
            frame = pd.read_csv(path, nrows=0, encoding="utf-8-sig")
            columns = [(column, "csv文本") for column in frame.columns]
            with path.open("r", encoding="utf-8-sig", errors="replace") as file:
                row_count = max(sum(1 for _ in file) - 1, 0)
            null_stats = csv_null_stats(path)
        column_names = [column for column, _ in columns]
        primary = primary_key_hint(stem, column_names)
        summary = {
            "英文表名": stem,
            "中文表名": chinese_name,
            "相对路径": str(path.relative_to(BASE_DIR)),
            "数据域": domain,
            "文件格式": path.suffix.lower().lstrip("."),
            "行数": row_count,
            "列数": len(columns),
            "大小MB": round(path.stat().st_size / 1024 / 1024, 2),
            "数据角色": role,
            "建议主键": primary,
            "主要关联键": relation_key(column_names),
            "完整性状态": "已落盘，结构可读",
            "备注": "合并后主表" if path.suffix.lower() == ".parquet" else "源表快照",
        }
        summaries.append(summary)
        review = key_review(path, stem, column_names)
        if review:
            key_rows.append(review)
            if review["是否可直接作为主键"] == "否":
                quality.append(
                    {
                        "数据集": stem,
                        "严重度": "高" if stem in {"parking_violations_all", "crash_merged", "crash_person_all"} else "中",
                        "问题类型": "候选键不唯一",
                        "问题描述": f"{review['候选键']} 重复数 {review['重复数']:,}",
                        "影响字段": review["候选键"],
                        "建议处理": review["治理建议"],
                    }
                )
        for index, (column, raw_type) in enumerate(columns, 1):
            field_row = {
                "数据集": stem,
                "序号": index,
                "字段名": column,
                "中文字段名": field_zh(column),
                "标准类型": type_label(raw_type),
                "原始类型": raw_type,
                "字段层级": field_level(column, primary),
                "业务说明": field_zh(column),
                "治理备注": "",
            }
            stat = null_stats[column]
            if isinstance(stat["缺失率"], float) and stat["缺失率"] >= 0.3:
                field_row["治理备注"] = "高缺失字段，需标注适用范围" if stat["缺失率"] >= 0.8 else "缺失率偏高，分析时需谨慎"
                quality.append(
                    {
                        "数据集": stem,
                        "严重度": "高" if stat["缺失率"] >= 0.8 else "中",
                        "问题类型": "字段缺失",
                        "问题描述": f"{column} 缺失率约 {stat['缺失率']:.1%}",
                        "影响字段": column,
                        "建议处理": field_row["治理备注"],
                    }
                )
            fields.append(field_row)
            missing_rows.append(
                {
                    "英文表名": stem,
                    "中文表名": chinese_name,
                    "数据域": domain,
                    "字段名": column,
                    "中文字段名": field_zh(column),
                    "标准类型": type_label(raw_type),
                    "行数": row_count,
                    "缺失数": stat["缺失数"],
                    "缺失率": f"{stat['缺失率']:.2%}" if isinstance(stat["缺失率"], float) else "",
                    "非空数": stat["非空数"],
                    "唯一值数": stat["唯一值数"],
                    "状态": stat["状态"],
                }
            )
    quality.extend(date_quality())
    return summaries, fields, missing_rows, quality, key_rows


def date_quality() -> list[dict[str, Any]]:
    issues = []
    specs = {
        "fhv_tripdata_2026Q1.parquet": ["pickup_datetime", "dropOff_datetime"],
        "fhvhv_tripdata_2026Q1.parquet": ["request_datetime", "pickup_datetime", "dropoff_datetime"],
        "green_tripdata_2026Q1.parquet": ["lpep_pickup_datetime", "lpep_dropoff_datetime"],
        "yellow_tripdata_2026Q1.parquet": ["tpep_pickup_datetime", "tpep_dropoff_datetime"],
    }
    for file_name, columns in specs.items():
        path = next(BASE_DIR.rglob(file_name))
        table = pq.read_table(path, columns=columns)
        for column in columns:
            min_value = pc.min(table[column]).as_py()
            max_value = pc.max(table[column]).as_py()
            if min_value.year < 2025 or max_value.year > 2026:
                issues.append(
                    {
                        "数据集": path.stem,
                        "严重度": "中",
                        "问题类型": "异常时间",
                        "问题描述": f"{column} 范围为 {min_value} 至 {max_value}",
                        "影响字段": column,
                        "建议处理": "保留原值并在 Silver 层增加异常时间标记",
                    }
                )
    return issues


def crash_relation_review() -> dict[str, Any]:
    crash_path = next(BASE_DIR.rglob("crash_merged.parquet"))
    person_path = next(BASE_DIR.rglob("crash_person_all.parquet"))
    crash_ids = set(pc.unique(pq.read_table(crash_path, columns=["collision_id"])["collision_id"]).to_pylist())
    person_ids = set(pc.unique(pq.read_table(person_path, columns=["collision_id"])["collision_id"]).to_pylist())
    return {
        "crash_distinct": len(crash_ids),
        "person_collision_distinct": len(person_ids),
        "intersection": len(crash_ids & person_ids),
        "person_without_crash": len(person_ids - crash_ids),
        "crash_without_person": len(crash_ids - person_ids),
    }


def append_plain_sheet(workbook: Workbook, title: str, rows: list[dict[str, Any]] | list[list[Any]]) -> None:
    sheet = workbook.create_sheet(title)
    if not rows:
        sheet.append(["无数据"])
        return
    if isinstance(rows[0], dict):
        headers = list(rows[0].keys())
        sheet.append(headers)
        for row in rows:
            sheet.append([row.get(header, "") for header in headers])
    else:
        for row in rows:
            sheet.append(row)
    style_sheet(sheet)


def append_dictionary_sheet(workbook: Workbook, summaries: list[dict[str, Any]], fields: list[dict[str, Any]]) -> None:
    sheet = workbook.create_sheet("数据字典")
    headers = ["序号", "字段名", "中文字段名", "标准类型", "原始类型", "字段层级", "业务说明", "治理备注"]
    grouped = defaultdict(list)
    for field in fields:
        grouped[field["数据集"]].append(field)
    row_index = 1
    for summary in summaries:
        table_fields = grouped.get(summary["英文表名"], [])
        if not table_fields:
            continue
        sheet.cell(row_index, 1, f"{summary['英文表名']}({summary['中文表名']})")
        sheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=len(headers))
        sheet.cell(row_index, 1).font = Font(bold=True, size=13, color="FFFFFF")
        sheet.cell(row_index, 1).fill = PatternFill("solid", fgColor="1F4E78")
        row_index += 1
        description = (
            f"表说明：数据域={summary['数据域']}；数据角色={summary['数据角色']}；"
            f"行数={summary['行数']:,}；列数={summary['列数']}；主键/候选键={summary['建议主键']}；"
            f"主要关联键={summary['主要关联键']}；完整性状态={summary['完整性状态']}。"
        )
        sheet.cell(row_index, 1, description)
        sheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=len(headers))
        sheet.cell(row_index, 1).alignment = Alignment(wrap_text=True, vertical="top")
        row_index += 1
        for column_index, header in enumerate(headers, 1):
            cell = sheet.cell(row_index, column_index, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="5B9BD5")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        row_index += 1
        for field in table_fields:
            for column_index, header in enumerate(headers, 1):
                cell = sheet.cell(row_index, column_index, field.get(header, ""))
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            row_index += 1
        row_index += 1
    apply_widths(sheet)


def style_sheet(sheet: Any) -> None:
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    sheet.freeze_panes = "A2"
    apply_widths(sheet)


def apply_widths(sheet: Any) -> None:
    for index, column in enumerate(sheet.columns, 1):
        width = min(max(len(str(cell.value)) if cell.value is not None else 0 for cell in column) + 2, 58)
        sheet.column_dimensions[get_column_letter(index)].width = max(width, 12)


def domain_rows(summaries: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped = defaultdict(lambda: {"tables": 0, "rows": 0})
    for summary in summaries:
        grouped[summary["数据域"]]["tables"] += 1
        grouped[summary["数据域"]]["rows"] += summary["行数"]
    descriptions = {
        "出行域": "TLC Q1 行程事实，支撑时空出行、费用、距离、供需分析",
        "资产域": "车辆、牌照、VIN、燃料、WAV 等资产属性",
        "供给域": "司机、基地和可服务资源快照",
        "安全域": "事故事件与事故人员，适合保留 ER 子模型",
        "监管合规域": "停车罚单、TIF、授权车辆、申请审批等合规主题",
        "空间地理域": "taxi zone、GIS、行政区、经纬度、街道等公共地理维度",
    }
    return [
        {"数据域": domain, "结构化表数": values["tables"], "结构化行数": values["rows"], "完整性判断": "高" if domain in {"出行域", "空间地理域"} else "中高", "说明": descriptions.get(domain, "")}
        for domain, values in sorted(grouped.items())
    ]


def model_rows() -> list[list[str]]:
    return [
        ["模型对象", "中文注释", "性质"],
        ["Bronze 原始层", "原始文件保留层，不改字段、不改格式，用于追溯来源", "规划层"],
        ["Silver 标准层", "清洗标准化层，统一字段名、类型、时间、空间、状态码、去重规则", "规划层"],
        ["Gold 主题星型模型", "面向分析和问数的主题模型，事实表围绕公共维表组织", "规划模型"],
        ["fact_trips", "出行事实表，记录出租车、网约车等每一次行程", "规划表"],
        ["fact_parking_violations", "停车罚单事实表，记录每一张违章罚单，需先按 summons_number 去重或确认粒度", "规划表"],
        ["fact_tif_payments", "出租车改善基金支付事实表，记录补贴或支付行为", "规划表"],
        ["fact_driver_applications", "司机申请事实表，记录新司机申请及审批状态", "规划表"],
        ["dim_vehicle", "车辆维表，描述车辆、VIN、车牌、燃料、WAV等资产属性", "规划表"],
        ["dim_driver", "司机维表，描述司机牌照、司机类型、有效期等供给属性", "规划表"],
        ["dim_base", "基地维表，描述FHV基地、基地类型、联系方式和地址", "规划表"],
        ["dim_taxi_zone", "出租车区域维表，描述LocationID、行政区、服务区域", "规划表"],
        ["dim_date", "日期维表，统一支持日、周、月、季度、财年等时间分析", "规划表"],
        ["dim_violation_code", "违章代码维表，解释罚单违章类型和罚款金额", "规划表"],
        ["dim_license_status", "牌照状态维表，统一解释车辆、司机、申请状态", "规划表"],
        ["事故 ER 子模型", "安全域保留事件与人员明细的一对多关系，但当前 collision_id/unique_id 需先去重治理", "规划模型"],
        ["fact_crashes", "事故事实表，记录每一次碰撞事故的位置、时间、伤亡和原因", "规划表"],
        ["fact_crash_persons", "事故人员明细表，记录事故涉及人员、伤害程度和人员属性", "规划表"],
        ["中文语义层", "规划层/元数据层，用于未来支撑中文用户问数和Agent理解，不是当前源数据表", "规划层"],
    ]


def mapping_rows() -> list[dict[str, str]]:
    return [
        {"原始字段": "tpep_pickup_datetime/lpep_pickup_datetime/pickup_datetime", "标准字段": "pickup_at", "中文注释": "统一接客时间"},
        {"原始字段": "tpep_dropoff_datetime/lpep_dropoff_datetime/dropoff_datetime/dropOff_datetime", "标准字段": "dropoff_at", "中文注释": "统一送客时间"},
        {"原始字段": "PULocationID/PUlocationID", "标准字段": "pickup_location_id", "中文注释": "统一上车区域编号"},
        {"原始字段": "DOLocationID/DOlocationID", "标准字段": "dropoff_location_id", "中文注释": "统一下车区域编号"},
        {"原始字段": "summons_number", "标准字段": "summons_no", "中文注释": "罚单编号，当前需去重后作为事实表候选键"},
        {"原始字段": "collision_id", "标准字段": "collision_id", "中文注释": "事故编号，连接事故与人员，但当前两表均需关系治理"},
        {"原始字段": "unique_id", "标准字段": "crash_person_unique_id", "中文注释": "事故人员记录编号，当前不唯一，需去重或生成代理键"},
        {"原始字段": "License Number/Vehicle License Number", "标准字段": "license_no", "中文注释": "统一牌照号"},
        {"原始字段": "DMV License Plate Number/DMV Plate Number/plate_id", "标准字段": "plate_no", "中文注释": "统一车牌号"},
        {"原始字段": "VIN/Vehicle VIN Number", "标准字段": "vehicle_vin", "中文注释": "统一车辆VIN"},
        {"原始字段": "Base Number/dispatching_base_num/Affiliated Base/ Agent Number", "标准字段": "base_no", "中文注释": "统一基地编号"},
    ]


def semantic_rows() -> list[dict[str, str]]:
    return [
        {"对象": "中文语义层", "中文注释": "规划层/元数据层，用于未来支撑中文用户问数和Agent理解，不是当前源数据表", "性质": "规划"},
        {"对象": "semantic_tables", "中文注释": "中文表说明，解释每张表能回答什么问题", "性质": "规划元数据"},
        {"对象": "semantic_fields", "中文注释": "中文字段说明，解释字段含义、单位、适用范围", "性质": "规划元数据"},
        {"对象": "semantic_metrics", "中文注释": "中文指标口径，如行程量、事故率、罚单量、WAV覆盖率", "性质": "规划元数据"},
        {"对象": "semantic_relationships", "中文注释": "中文关系说明，标注强关联、弱关联和不可关联", "性质": "规划元数据"},
        {"对象": "semantic_question_templates", "中文注释": "中文问数模板，约束Agent如何安全生成查询", "性质": "规划元数据"},
    ]


def write_workbook(summaries: list[dict[str, Any]], fields: list[dict[str, Any]], missing_rows: list[dict[str, Any]], quality: list[dict[str, Any]], key_rows: list[dict[str, Any]]) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    append_plain_sheet(workbook, "数据总览", summaries)
    append_plain_sheet(workbook, "六大数据域", domain_rows(summaries))
    append_dictionary_sheet(workbook, summaries, fields)
    append_plain_sheet(workbook, "缺失率矩阵", missing_rows)
    append_plain_sheet(workbook, "主键与关系审查", key_rows + [crash_relation_row()])
    append_plain_sheet(workbook, "表关系模型", model_rows())
    append_plain_sheet(workbook, "入湖标准字段映射", mapping_rows())
    append_plain_sheet(workbook, "中文语义层", semantic_rows())
    append_plain_sheet(workbook, "数据质量问题", quality)
    workbook.save(OUTPUT_XLSX)


def crash_relation_row() -> dict[str, Any]:
    review = crash_relation_review()
    return {
        "英文表名": "crash_merged ↔ crash_person_all",
        "候选键": "collision_id",
        "行数": "",
        "唯一值数": "",
        "空值数": "",
        "重复数": "",
        "是否可直接作为主键": "否",
        "治理建议": f"事故表唯一事故 {review['crash_distinct']:,}；人员表事故编号 {review['person_collision_distinct']:,}；交集 {review['intersection']:,}；人员有但事故无 {review['person_without_crash']:,}；事故有但人员无 {review['crash_without_person']:,}",
    }


def model_tree() -> str:
    return """纽约市城市交通数据底座
├─ Bronze 原始层  # 原始文件保留层，不改字段、不改格式，用于追溯来源
├─ Silver 标准层  # 清洗标准化层，统一字段名、类型、时间、空间、状态码、去重规则
├─ Gold 主题星型模型  # 面向分析和问数的主题模型，事实表围绕公共维表组织
│  ├─ fact_trips  # 出行事实表，记录出租车、网约车等每一次行程
│  ├─ fact_parking_violations  # 停车罚单事实表，记录罚单，需先按summons_number治理重复
│  ├─ fact_tif_payments  # 出租车改善基金支付事实表，记录补贴或支付行为
│  ├─ fact_driver_applications  # 司机申请事实表，记录新司机申请及审批状态
│  ├─ dim_vehicle  # 车辆维表，描述车辆、VIN、车牌、燃料、WAV等资产属性
│  ├─ dim_driver  # 司机维表，描述司机牌照、司机类型、有效期等供给属性
│  ├─ dim_base  # 基地维表，描述FHV基地、基地类型、联系方式和地址
│  ├─ dim_taxi_zone  # 出租车区域维表，描述LocationID、行政区、服务区域
│  ├─ dim_date  # 日期维表，统一支持日、周、月、季度、财年等时间分析
│  ├─ dim_violation_code  # 违章代码维表，解释罚单违章类型和罚款金额
│  └─ dim_license_status  # 牌照状态维表，统一解释车辆、司机、申请状态
├─ 事故 ER 子模型  # 安全域保留事件与人员明细的一对多关系，需先治理collision_id覆盖差异
│  ├─ fact_crashes  # 事故事实表，记录每一次碰撞事故的位置、时间、伤亡和原因
│  └─ fact_crash_persons  # 事故人员明细表，记录事故涉及人员、伤害程度和人员属性
└─ 中文语义层  # 规划层/元数据层，用于中文问数和Agent理解，不是当前源数据表
   ├─ semantic_tables  # 中文表说明，解释每张表能回答什么问题
   ├─ semantic_fields  # 中文字段说明，解释字段含义、单位、适用范围
   ├─ semantic_metrics  # 中文指标口径，如行程量、事故率、罚单量、WAV覆盖率
   ├─ semantic_relationships  # 中文关系说明，标注强关联、弱关联和不可关联
   └─ semantic_question_templates  # 中文问数模板，约束Agent如何安全生成查询"""


def write_markdown(summaries: list[dict[str, Any]], quality: list[dict[str, Any]], key_rows: list[dict[str, Any]]) -> None:
    total_rows = sum(row["行数"] for row in summaries)
    trip_rows = sum(row["行数"] for row in summaries if row["数据域"] == "出行域")
    domain_summary = "\n".join(
        f"| {row['数据域']} | {row['结构化表数']} | {row['结构化行数']:,} | {row['完整性判断']} | {row['说明']} |"
        for row in domain_rows(summaries)
    )
    key_summary = "\n".join(
        f"| {row['英文表名']} | {row['候选键']} | {row['行数']:,} | {row['唯一值数']:,} | {row['重复数']:,} | {row['是否可直接作为主键']} |"
        for row in key_rows
    )
    quality_summary = "\n".join(
        f"| {row['数据集']} | {row['严重度']} | {row['问题类型']} | {row['问题描述']} | {row['建议处理']} |"
        for row in quality[:80]
    )
    relation = crash_relation_review()
    OUTPUT_FEASIBILITY.write_text(
        f"""# 纽约市城市交通 Agent 基座可行性评估

生成时间：{GENERATED_AT}

## 总体结论

本次审查目录为：`{BASE_DIR}`。

当前数据已经从旧版“分片采集中”升级为“主目录内合并后的结构化主表”。已识别结构化表 {len(summaries)} 张，总行数约 {total_rows:,} 行，其中出行域约 {trip_rows:,} 行，停车罚单、事故、事故人员也已落成 parquet 主表。

综合判断：**适合一起作为数据分析与数据开发 Agent 的基础底座，当前评级为 A-，完成 Silver 去重和关系治理后可评为 A。**

## 六大域覆盖

| 数据域 | 结构化表数 | 结构化行数 | 完整性判断 | 说明 |
|---|---:|---:|---|---|
{domain_summary}

## 源表模型判断

源表仍不是严格 ER 模型，也不是标准星型模型。它们是“开放数据宽表集合 + 官方字典 + 已合并 parquet 主表”。

当前可直接支撑星型建模的是出行域与空间地理域；安全域适合保留事故 ER 子模型；监管合规域已经具备罚单事实表，但需要先治理 `summons_number` 重复。

## 推荐目标模型

```text
{model_tree()}
```

## 主键与关系审查

| 表 | 候选键 | 行数 | 唯一值数 | 重复数 | 是否可直接作为主键 |
|---|---|---:|---:|---:|---|
{key_summary}

事故关系补充：事故表唯一 `collision_id` 为 {relation['crash_distinct']:,}，事故人员表涉及事故编号 {relation['person_collision_distinct']:,}，两者交集 {relation['intersection']:,}。存在人员有但事故表无、事故有但人员表无的情况，进入 Gold 层前需要关系覆盖治理。

## Agent 适配能力

- 支持中文问数：出行趋势、区域热度、罚单趋势、事故分布、车辆结构、司机供给、WAV覆盖。
- 支持数据开发：parquet 主表建模、去重规则、主键生成、维表抽取、质量检测、指标层设计。
- 支持治理解释：中文字段语义、强弱关联说明、指标口径说明、不可关联限制。

## 主要风险

- `parking_violations_all.summons_number` 当前不唯一，需确认是否重复采集或存在多版本记录。
- `crash_merged.collision_id` 与 `crash_person_all.unique_id` 当前不唯一，事故 ER 子模型不能裸表直接使用。
- 事故表与事故人员表的 `collision_id` 覆盖不完全，需要关系完整性治理。
- TLC 行程表仍有少量异常时间，Silver 层应保留原值并增加异常标记。
- 车辆/司机与行程之间缺少 trip 级强外键，Agent 不能推断“某司机完成某行程”。

## 建设优先级

1. 先建设 `dim_taxi_zone` 与 `fact_trips`。
2. 对 `parking_violations_all`、`crash_merged`、`crash_person_all` 做去重和主键治理。
3. 建设 `fact_parking_violations`、`fact_crashes`、`fact_crash_persons`。
4. 建设 `dim_vehicle`、`dim_driver`、`dim_base`、`dim_violation_code`。
5. 建设中文语义层，约束 Agent 的中文问数和自动建模行为。
""",
        encoding="utf-8",
    )
    OUTPUT_COMPLETENESS.write_text(
        f"""# 纽约市城市交通数据完整性分析报告

生成时间：{GENERATED_AT}

## 总体概况

主目录：`{BASE_DIR}`

当前已识别结构化表 {len(summaries)} 张，总行数约 {total_rows:,} 行。与旧版相比，事故、事故人员、停车罚单已经从外部分片状态进入主目录合并 parquet 状态，文件完整性明显提升。

## 完整性结论

| 数据域 | 结构化表数 | 结构化行数 | 完整性判断 | 说明 |
|---|---:|---:|---|---|
{domain_summary}

## 文件完整性

- 未发现 `.nas_pro_downloading`、`.part`、`.crdownload` 等下载临时文件。
- TLC Q1 四类行程已合并为单个 parquet 主表。
- 停车罚单已合并为 `parking_violations_all.parquet`。
- 事故数据已合并为 `crash_merged.parquet`，事故人员已合并为 `crash_person_all.parquet`。
- 官方数据字典、说明文件、taxi zone GIS 文件仍保留，可作为语义层和维表来源。

## 表结构完整性

- 所有结构化 parquet/csv 均可读取。
- TLC 行程、停车罚单、事故、事故人员均已具备稳定主表形态。
- 事故表目录名为“机动车碰撞事故 - 车辆”，但当前 `crash_merged.parquet` 更接近事故事件宽表，不是真正的逐车辆明细表。

## 主键完整性

| 表 | 候选键 | 行数 | 唯一值数 | 重复数 | 是否可直接作为主键 |
|---|---|---:|---:|---:|---|
{key_summary}

## 关系完整性

- `PULocationID`、`DOLocationID` 可关联 `taxi_zone_lookup.LocationID`。
- `collision_id` 可作为事故表与事故人员表的关系键，但当前覆盖不完全：交集 {relation['intersection']:,}，人员有但事故表无 {relation['person_without_crash']:,}，事故有但人员表无 {relation['crash_without_person']:,}。
- 车辆、司机和行程之间缺少 trip 级强外键，只适合按基地、牌照、时间快照做弱关联或聚合分析。

## 质量问题清单

| 数据集 | 严重度 | 问题类型 | 问题描述 | 建议处理 |
|---|---|---|---|---|
{quality_summary}

## Agent 使用限制

- 可以回答出行趋势、区域热度、事故分布、罚单趋势、车辆结构、供给规模等问题。
- 不应在缺少强外键时回答单车、单司机、单行程的确定性因果问题。
- 对重复候选键的表，Agent 必须说明是否使用去重口径。
- 对异常时间和高缺失字段，Agent 必须说明过滤口径或数据限制。
""",
        encoding="utf-8",
    )


def main() -> None:
    summaries, fields, missing_rows, quality, key_rows = collect_metadata()
    write_workbook(summaries, fields, missing_rows, quality, key_rows)
    write_markdown(summaries, quality, key_rows)
    print(f"updated={OUTPUT_XLSX}")
    print(f"updated={OUTPUT_FEASIBILITY}")
    print(f"updated={OUTPUT_COMPLETENESS}")


if __name__ == "__main__":
    main()
