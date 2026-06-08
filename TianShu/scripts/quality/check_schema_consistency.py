"""
数据库设计文档、字段字典和 DuckDB schema 一致性检查

当前 Silver 尚未建表，因此本脚本先检查：
1. 数据库设计文档是否存在
2. Bronze 设计文档字段数是否与 DuckDB 实际 schema 一致
3. Silver 数据库设计文档字段数是否与 Silver xlsx 一致
4. Silver xlsx 总览字段数是否与各表 sheet 字段数一致
"""
import argparse
import re
import sys
from pathlib import Path

try:
    import duckdb
except ImportError:
    duckdb = None

try:
    import openpyxl
except ImportError:
    openpyxl = None


def read_text(path: Path) -> str:
    """读取 UTF-8 文档"""
    return path.read_text(encoding="utf-8")


def parse_markdown_table(content: str) -> list[list[str]]:
    """解析简单 Markdown 表格"""
    rows: list[list[str]] = []
    for line in content.splitlines():
        line = line.strip()
        if not line.startswith("|") or "---" in line:
            continue
        cells = [c.strip().strip("`") for c in line.strip("|").split("|")]
        if cells:
            rows.append(cells)
    return rows


def load_bronze_field_counts(db_path: Path) -> dict[str, int]:
    """读取 DuckDB 中 Bronze 层实际字段数"""
    if duckdb is None:
        raise RuntimeError("duckdb 未安装，无法检查实际 schema")
    conn = duckdb.connect(str(db_path), read_only=True)
    try:
        rows = conn.execute("""
            SELECT table_name, count(*) AS field_count
            FROM information_schema.columns
            WHERE table_schema = 'bronze'
            GROUP BY table_name
        """).fetchall()
        return {table: count for table, count in rows}
    finally:
        conn.close()


def load_silver_xlsx_counts(xlsx_path: Path) -> tuple[dict[str, int], dict[str, int]]:
    """读取 Silver xlsx 总览字段数和各 sheet 实际字段数"""
    if openpyxl is None:
        raise RuntimeError("openpyxl 未安装，无法读取 xlsx")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        overview_counts: dict[str, int] = {}
        sheet_counts: dict[str, int] = {}

        ws = wb["Silver表清单"]
        headers = [c.value for c in next(ws.iter_rows(max_row=1))]
        table_idx = headers.index("英文表名")
        count_idx = headers.index("字段数")
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[table_idx]:
                continue
            table_name = str(row[table_idx]).replace("silver.", "")
            overview_counts[table_name] = int(row[count_idx])

        for sheet_name in wb.sheetnames:
            if sheet_name == "Silver表清单":
                continue
            ws_sheet = wb[sheet_name]
            sheet_counts[sheet_name] = max(ws_sheet.max_row - 1, 0)

        return overview_counts, sheet_counts
    finally:
        wb.close()


def parse_design_field_counts(design_path: Path) -> dict[str, int]:
    """从 Silver 数据库设计文档表清单解析字段数"""
    rows = parse_markdown_table(read_text(design_path))
    counts: dict[str, int] = {}
    for cells in rows:
        if len(cells) < 6 or cells[1] == "英文表名":
            continue
        table_name = cells[1].replace("silver.", "")
        field_count = cells[5]
        if re.fullmatch(r"\d+", field_count):
            counts[table_name] = int(field_count)
    return counts


def parse_bronze_design_field_counts(design_path: Path) -> dict[str, int]:
    """从 Bronze 数据库设计文档表清单解析字段数"""
    rows = parse_markdown_table(read_text(design_path))
    counts: dict[str, int] = {}
    for cells in rows:
        if len(cells) < 7 or cells[0] == "英文表名":
            continue
        table_name = cells[0]
        field_count = cells[6]
        if re.fullmatch(r"\d+", field_count):
            counts[table_name] = int(field_count)
    return counts


def check_required_docs(project_root: Path) -> list[str]:
    """检查最高事实源文档是否存在"""
    required = [
        "docs/warehouse/database_design/README.md",
        "docs/warehouse/database_design/bronze_database_design.md",
        "docs/warehouse/database_design/silver_database_design.md",
        "docs/warehouse/database_design/gold_database_design.md",
        "docs/warehouse/data_dictionary/README.md",
        "docs/memory/经验复盘.md",
        "docs/memory/风险清单.md",
        "docs/memory/规则来源索引.md",
    ]
    violations: list[str] = []
    for rel in required:
        if not (project_root / rel).exists():
            violations.append(f"缺少必要文档: {rel}")
    return violations


def compare_counts(label: str, left: dict[str, int], right: dict[str, int]) -> list[str]:
    """比较两组表字段数"""
    violations: list[str] = []
    for table_name, left_count in left.items():
        if table_name not in right:
            violations.append(f"[{label}] {table_name} 只存在于左侧，右侧缺失")
            continue
        right_count = right[table_name]
        if left_count != right_count:
            violations.append(
                f"[{label}] {table_name} 字段数不一致: 左侧={left_count}, 右侧={right_count}"
            )
    for table_name in right:
        if table_name not in left:
            violations.append(f"[{label}] {table_name} 只存在于右侧，左侧缺失")
    return violations


def main() -> int:
    parser = argparse.ArgumentParser(description="schema 一致性检查")
    parser.add_argument("--project-root", default=r"D:\Program Files\gitvscode\TianShu")
    parser.add_argument("--db", default=r"D:\ProgramData\Datawarehouse\纽约市城市交通\nyc_transport.duckdb")
    parser.add_argument("--silver-xlsx", default=r"D:\ProgramData\Datawarehouse\纽约市城市交通\分析报告\Silver层数据字典.xlsx")
    args = parser.parse_args()

    project_root = Path(args.project_root)
    db_path = Path(args.db)
    xlsx_path = Path(args.silver_xlsx)

    violations: list[str] = []
    warnings: list[str] = []

    print("=" * 60)
    print("schema 一致性检查")
    print("=" * 60)

    violations.extend(check_required_docs(project_root))

    if db_path.exists():
        bronze_actual = load_bronze_field_counts(db_path)
        bronze_design = parse_bronze_design_field_counts(
            project_root / "docs/warehouse/database_design/bronze_database_design.md"
        )
        violations.extend(compare_counts("Bronze设计文档 vs DuckDB", bronze_design, bronze_actual))
    else:
        violations.append(f"DuckDB 数据库不存在: {db_path}")

    if xlsx_path.exists():
        overview_counts, sheet_counts = load_silver_xlsx_counts(xlsx_path)
        silver_design = parse_design_field_counts(
            project_root / "docs/warehouse/database_design/silver_database_design.md"
        )
        violations.extend(compare_counts("Silver xlsx总览 vs sheet", overview_counts, sheet_counts))
        violations.extend(compare_counts("Silver设计文档 vs xlsx总览", silver_design, overview_counts))
    else:
        violations.append(f"Silver xlsx 不存在: {xlsx_path}")

    warnings.append("Silver 实表尚未建设，已跳过 silver schema 与设计文档一致性检查。")

    for warning in warnings:
        print(f"[WARN] {warning}")
    for violation in violations:
        print(f"[FAIL] {violation}")

    print("=" * 60)
    if violations:
        print(f"[FAIL] 共发现 {len(violations)} 个一致性问题。")
        return 1
    print("[OK] schema 一致性检查通过。")
    return 0


if __name__ == "__main__":
    sys.exit(main())
