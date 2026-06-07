"""
Silver 数据字典一致性检查

检查项：
1. Silver xlsx中的每个"直接来源"字段是否存在于Bronze DESCRIBE结果
2. 派生字段是否填写了来源和计算逻辑
3. 是否出现已知的危险字段名（如Bronze中不存在的金额字段）
4. 字段数是否与Markdown规划文档一致

用法：python check_silver_dictionary.py [--xlsx <路径>] [--bronze-db <路径>]
"""
import argparse
import sys
import re
from pathlib import Path
from typing import Optional

try:
    import duckdb
except ImportError:
    duckdb = None

try:
    import openpyxl
except ImportError:
    openpyxl = None


# 已知Bronze中不存在的字段（经验复盘沉淀）
# 格式：{表名: {字段名集合}}
FORBIDDEN_DIRECT_FIELDS: dict[str, set[str]] = {
    "parking_violation_detail": {
        "fine_amount", "penalty_amount", "interest_amount",
        "reduction_amount", "payment_amount", "amount_due",
    },
}

# 已知危险字段名模式（跨表通用）
DANGEROUS_FIELD_PATTERNS: list[tuple[str, str]] = [
    (r"^(fine|penalty|interest|reduction|payment)_amount$", "金额字段"),
    (r"_amount$", "金额字段"),
    (r"^amount_", "金额字段"),
]


def load_bronze_columns(db_path: str) -> dict[str, set[str]]:
    """从 DuckDB 的 information_schema 加载 Bronze 层所有字段"""
    if duckdb is None:
        print("[ERROR] duckdb 未安装，无法连接数据库。请 pip install duckdb")
        sys.exit(1)
    conn = duckdb.connect(db_path, read_only=True)
    try:
        result = conn.execute("""
            SELECT table_name, column_name
            FROM information_schema.columns
            WHERE table_schema = 'bronze'
            ORDER BY table_name, ordinal_position
        """).fetchall()
        bronze_cols: dict[str, set[str]] = {}
        for table, col in result:
            bronze_cols.setdefault(table, set()).add(col)
        return bronze_cols
    finally:
        conn.close()


def load_silver_xlsx(xlsx_path: str) -> dict[str, list[dict]]:
    """加载 Silver 数据字典 xlsx，返回 {sheet名: [字段行dict]}"""
    if openpyxl is None:
        print("[ERROR] openpyxl 未安装。请 pip install openpyxl")
        sys.exit(1)
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    result: dict[str, list[dict]] = {}
    for sheet_name in wb.sheetnames:
        if sheet_name == "Silver表清单":
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(h).strip() if h else "" for h in rows[0]]
        fields: list[dict] = []
        for row in rows[1:]:
            field = {}
            for i, h in enumerate(headers):
                if i < len(row) and row[i] is not None:
                    field[h] = str(row[i]).strip()
            if field.get("英文字段名"):
                fields.append(field)
        result[sheet_name] = fields
    wb.close()
    return result


def split_source_items(value: str) -> list[str]:
    """拆分字段字典中可能包含多个来源字段的文本"""
    if not value:
        return []
    cleaned = (
        value.replace("`", "")
        .replace("，", ",")
        .replace("、", ",")
        .replace("/", ",")
        .replace("；", ";")
        .replace(";", ",")
    )
    result: list[str] = []
    for item in cleaned.split(","):
        item = item.strip()
        if not item:
            continue
        item = re.split(r"[（(]", item, maxsplit=1)[0].strip()
        if item and item not in ("无", "详见单表规划文档"):
            result.append(item)
    return result


def normalize_source_tables(value: str | list[str]) -> list[str]:
    """统一把来源表配置转成列表"""
    if isinstance(value, list):
        return value
    if not value:
        return []
    return [v.strip() for v in value.split(",") if v.strip()]


def check_source_fields(
    silver_dict: dict[str, list[dict]],
    bronze_cols: dict[str, set[str]],
    table_bronze_map: dict[str, str | list[str]],
) -> list[str]:
    """检查 direct/standardized 字段的来源是否存在于 Bronze"""
    violations: list[str] = []
    for table_name, fields in silver_dict.items():
        bronze_tables = normalize_source_tables(table_bronze_map.get(table_name, ""))
        bronze_sets = {t: bronze_cols.get(t, set()) for t in bronze_tables}

        for f in fields:
            field_en = f.get("英文字段名", "")
            source_type = f.get("字段来源类型", f.get("source_type", ""))
            source_col = f.get("来源Bronze字段", f.get("来源字段", ""))
            derivation = f.get("派生逻辑", f.get("计算逻辑", ""))

            # 检查禁止字段
            forbidden = FORBIDDEN_DIRECT_FIELDS.get(table_name, set())
            if field_en in forbidden and source_type != "derived":
                violations.append(
                    f"[{table_name}] 字段 '{field_en}' 在 Bronze 中不存在，"
                    f"曾被确认为虚假字段（经验复盘 2026-06-07）。"
                    f"如果确实需要此字段，必须标注为 'derived' 且填写完整的派生逻辑。"
                )

            if source_type in ("direct", "standardized"):
                source_items = split_source_items(source_col)
                if source_items and bronze_sets:
                    found = any(
                        col in cols
                        for col in source_items
                        for cols in bronze_sets.values()
                    )
                    if found:
                        continue
                    violations.append(
                        f"[{table_name}] 字段 '{field_en}' 标注为 {source_type}，"
                        f"声称来源于 '{source_col}'，"
                        f"但来源 Bronze 表 {bronze_tables} 的 DESCRIBE 结果中均不存在这些列。"
                    )
            elif source_type == "derived":
                if not derivation:
                    violations.append(
                        f"[{table_name}] 字段 '{field_en}' 标注为 derived，"
                        f"但未填写派生逻辑。"
                    )

    return violations


def check_dangerous_patterns(
    silver_dict: dict[str, list[dict]],
) -> list[str]:
    """扫描是否有已知危险字段模式"""
    violations: list[str] = []
    for table_name, fields in silver_dict.items():
        for f in fields:
            field_en = f.get("英文字段名", "")
            source_type = f.get("字段来源类型", f.get("source_type", ""))
            derivation = f.get("派生逻辑", "")

            for pattern, category in DANGEROUS_FIELD_PATTERNS:
                if re.match(pattern, field_en):
                    if not source_type:
                        violations.append(
                            f"[{table_name}] 字段 '{field_en}' 匹配危险模式 "
                            f"'{category}'，但未标注 source_type。"
                            f"请确认此字段是否存在于 Bronze DESCRIBE 结果。"
                        )
                    elif source_type == "derived" and not derivation:
                        violations.append(
                            f"[{table_name}] 字段 '{field_en}' 为 derived {category}，"
                            f"但未填写派生逻辑。"
                        )
    return violations


def check_field_count_consistency(
    silver_dict: dict[str, list[dict]],
    plan_dir: str,
) -> list[str]:
    """检查 xlsx 字段数与 Markdown 规划文档声明的字段数是否一致"""
    violations: list[str] = []
    for table_name, fields in silver_dict.items():
        xlsx_count = len(fields)
        md_path = Path(plan_dir) / f"{table_name.replace('silver.', '')}.md"
        if not md_path.exists():
            continue
        content = md_path.read_text(encoding="utf-8")
        m = re.search(r"字段数[：:]\s*(\d+)", content)
        if m:
            md_count = int(m.group(1))
            if xlsx_count != md_count:
                violations.append(
                    f"[{table_name}] xlsx 字段数({xlsx_count}) "
                    f"≠ MD 声明字段数({md_count})"
                )
    return violations


def main():
    parser = argparse.ArgumentParser(description="Silver 数据字典一致性检查")
    parser.add_argument(
        "--xlsx",
        default=r"D:\ProgramData\Datawarehouse\纽约市城市交通\分析报告\Silver层数据字典.xlsx",
        help="Silver 数据字典 xlsx 路径",
    )
    parser.add_argument(
        "--bronze-db",
        default=r"D:\ProgramData\Datawarehouse\纽约市城市交通\nyc_transport.duckdb",
        help="DuckDB 数据库路径",
    )
    parser.add_argument(
        "--plan-dir",
        default=r"D:\Program Files\gitvscode\TianShu\scripts\silver",
        help="Silver 规划文档目录",
    )
    args = parser.parse_args()

    if not Path(args.xlsx).exists():
        print(f"[SKIP] xlsx 文件不存在: {args.xlsx}")
        print("请先运行 scripts/silver/_gen_xlsx.py 生成数据字典")
        return

    bronze_cols = load_bronze_columns(args.bronze_db)
    silver_dict = load_silver_xlsx(args.xlsx)

    # Silver 表名 → Bronze 来源表名 映射
    table_bronze_map = {
        "trip_detail": [
            "yellow_tripdata_2026q1",
            "green_tripdata_2026q1",
            "fhv_tripdata_2026q1",
            "fhvhv_tripdata_2026q1",
        ],
        "taxi_zone": "taxi_zone_lookup",
        "dim_date": "",  # 全部派生，无直接Bronze来源
        "vehicle_detail": [
            "active_vehicles",
            "fhv_active_vehicles",
            "medallion_authorized_vehicles",
        ],
        "driver_detail": [
            "fhv_active_drivers",
            "shl_active_drivers",
        ],
        "base_detail": "fhv_base_aggregate_report",
        "driver_application_detail": "new_driver_applications",
        "parking_violation_detail": "parking_violations_all",
        "tif_payment_detail": "tif_medallion_payments",
        "crash_detail": "crash_merged",
        "crash_person_detail": "crash_person_all",
    }

    all_violations: list[str] = []

    print("=" * 60)
    print("Silver 数据字典一致性检查")
    print("=" * 60)

    # 检查1：字段来源
    v1 = check_source_fields(silver_dict, bronze_cols, table_bronze_map)
    all_violations.extend(v1)
    print(f"\n[1] 字段来源检查: {len(v1)} 个违规")
    for v in v1:
            print(f"  [FAIL] {v}")

    # 检查2：危险模式
    v2 = check_dangerous_patterns(silver_dict)
    all_violations.extend(v2)
    print(f"\n[2] 危险模式扫描: {len(v2)} 个违规")
    for v in v2:
            print(f"  [WARN] {v}")

    # 检查3：字段数一致性
    v3 = check_field_count_consistency(silver_dict, args.plan_dir)
    all_violations.extend(v3)
    print(f"\n[3] 字段数一致性: {len(v3)} 个违规")
    for v in v3:
            print(f"  [FAIL] {v}")

    print(f"\n{'=' * 60}")
    if all_violations:
        print(f"[FAIL] 共发现 {len(all_violations)} 个违规。请修复后重新检查。")
        sys.exit(1)
    else:
        print("[OK] 全部检查通过。")
        sys.exit(0)


if __name__ == "__main__":
    main()
